import json
from datetime import datetime, timedelta
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.db.models import Q, F, Sum, Count
from django.db import transaction
from openpyxl import Workbook
from .models import Producto, MovimientoStock, VentaBarra, Asistente
from .decorators import barra_o_admin, solo_barra, ajax_login_required

# ===== DASHBOARD BARRA =====

@barra_o_admin
def dashboard_barra(request):
    """Dashboard específico para personal de barra"""
    hoy = datetime.now().date()
    
    # Estadísticas del día
    stats = {
        'productos_activos': Producto.objects.filter(activo=True).count(),
        'productos_bajo_stock': Producto.objects.filter(stock__lte=F('stock_minimo'), activo=True).count(),
        'productos_agotados': Producto.objects.filter(stock=0, activo=True).count(),
        'mis_ventas_hoy': VentaBarra.objects.filter(vendedor=request.user, fecha__date=hoy).count(),
        'total_vendido_hoy': VentaBarra.objects.filter(
            vendedor=request.user, fecha__date=hoy
        ).aggregate(total=Sum('total'))['total'] or 0,
        'ventas_totales_hoy': VentaBarra.objects.filter(fecha__date=hoy).count(),
        'ingresos_totales_hoy': VentaBarra.objects.filter(
            fecha__date=hoy
        ).aggregate(total=Sum('total'))['total'] or 0
    }
    
    # Productos críticos (bajo stock)
    productos_criticos = Producto.objects.filter(
        stock__lte=F('stock_minimo'), activo=True
    ).order_by('stock')[:5]
    
    # Productos más vendidos hoy
    productos_top = Producto.objects.filter(
        ventabarra__fecha__date=hoy
    ).annotate(
        ventas_hoy=Count('ventabarra'),
        total_vendido=Sum('ventabarra__total')
    ).order_by('-ventas_hoy')[:5]
    
    # Últimas ventas del usuario
    ultimas_ventas = VentaBarra.objects.filter(
        vendedor=request.user, fecha__date=hoy
    ).order_by('-fecha')[:5]
    
    context = {
        'stats': stats,
        'productos_criticos': productos_criticos,
        'productos_top': productos_top,
        'ultimas_ventas': ultimas_ventas,
    }
    
    return render(request, 'core/barra/dashboard.html', context)

# ===== CONTROL DE INVENTARIO =====

@barra_o_admin
def inventario(request):
    """Vista general del inventario"""
    productos = Producto.objects.filter(activo=True).order_by('nombre')
    
    # Filtros
    buscar = request.GET.get('buscar', '').strip()
    if buscar:
        productos = productos.filter(nombre__icontains=buscar)
    
    estado = request.GET.get('estado')
    if estado == 'bajo_stock':
        productos = productos.filter(stock__lte=F('stock_minimo'))
    elif estado == 'agotado':
        productos = productos.filter(stock=0)
    elif estado == 'disponible':
        productos = productos.filter(stock__gt=0)
    
    # Estadísticas
    stats = {
        'total_productos': productos.count(),
        'valor_inventario': sum(p.precio * p.stock for p in productos),
        'productos_bajo_stock': productos.filter(stock__lte=F('stock_minimo')).count(),
        'productos_agotados': productos.filter(stock=0).count()
    }
    
    context = {
        'productos': productos,
        'stats': stats,
        'filtros': {
            'buscar': buscar,
            'estado': estado
        }
    }
    
    return render(request, 'core/barra/inventario.html', context)

@barra_o_admin
def lista_productos(request):
    """Lista completa de productos"""
    productos = Producto.objects.filter(activo=True).order_by('nombre')
    
    context = {'productos': productos}
    return render(request, 'core/barra/productos_lista.html', context)

@barra_o_admin
def crear_producto(request):
    """Crear nuevo producto"""
    if request.method == 'POST':
        try:
            with transaction.atomic():
                producto = Producto.objects.create(
                    nombre=request.POST['nombre'].strip(),
                    descripcion=request.POST.get('descripcion', '').strip(),
                    precio=float(request.POST['precio']),
                    stock=int(request.POST['stock']),
                    stock_minimo=int(request.POST.get('stock_minimo', 5)),
                    creado_por=request.user
                )
                
                # Crear movimiento de stock inicial
                if producto.stock > 0:
                    MovimientoStock.objects.create(
                        producto=producto,
                        tipo='entrada',
                        cantidad=producto.stock,
                        observacion='Stock inicial',
                        usuario=request.user
                    )
                
                messages.success(request, f'Producto {producto.nombre} creado exitosamente')
                return redirect('barra:lista_productos')
        except Exception as e:
            messages.error(request, f'Error al crear producto: {e}')
    
    return render(request, 'core/barra/producto_form.html')

@barra_o_admin
def editar_producto(request, pk):
    """Editar producto existente"""
    producto = get_object_or_404(Producto, pk=pk, activo=True)
    
    if request.method == 'POST':
        try:
            producto.nombre = request.POST['nombre'].strip()
            producto.descripcion = request.POST.get('descripcion', '').strip()
            producto.precio = float(request.POST['precio'])
            producto.stock_minimo = int(request.POST.get('stock_minimo', 5))
            producto.save()
            
            messages.success(request, f'Producto {producto.nombre} actualizado exitosamente')
            return redirect('barra:lista_productos')
        except Exception as e:
            messages.error(request, f'Error al actualizar producto: {e}')
    
    context = {'producto': producto, 'editando': True}
    return render(request, 'core/barra/producto_form.html', context)

@barra_o_admin
def ajustar_stock(request, pk):
    """Ajustar stock de un producto"""
    producto = get_object_or_404(Producto, pk=pk, activo=True)
    
    if request.method == 'POST':
        try:
            nuevo_stock = int(request.POST['nuevo_stock'])
            observacion = request.POST.get('observacion', '').strip()
            
            with transaction.atomic():
                MovimientoStock.objects.create(
                    producto=producto,
                    tipo='ajuste',
                    cantidad=nuevo_stock,
                    observacion=f'Ajuste: {observacion}' if observacion else 'Ajuste de stock',
                    usuario=request.user
                )
            
            messages.success(request, f'Stock ajustado para {producto.nombre}')
            return redirect('barra:inventario')
        except Exception as e:
            messages.error(request, f'Error: {e}')
    
    return render(request, 'core/barra/ajustar_stock.html', {'producto': producto})

# ===== MOVIMIENTOS DE STOCK =====

@barra_o_admin
def movimientos_stock(request):
    """Ver historial de movimientos de stock"""
    movimientos = MovimientoStock.objects.all().order_by('-fecha')[:100]
    productos = Producto.objects.filter(activo=True).order_by('nombre')
    
    # Filtros
    producto_id = request.GET.get('producto')
    if producto_id:
        movimientos = movimientos.filter(producto_id=producto_id)
    
    tipo = request.GET.get('tipo')
    if tipo:
        movimientos = movimientos.filter(tipo=tipo)
    
    context = {
        'movimientos': movimientos,
        'productos': productos,
        'tipos': MovimientoStock.TIPOS,
        'filtros': {
            'producto': producto_id,
            'tipo': tipo
        }
    }
    
    return render(request, 'core/barra/movimientos_stock.html', context)

@barra_o_admin
def registrar_movimiento(request):
    """Registrar nuevo movimiento de stock"""
    if request.method == 'POST':
        try:
            with transaction.atomic():
                MovimientoStock.objects.create(
                    producto_id=request.POST['producto'],
                    tipo=request.POST['tipo'],
                    cantidad=int(request.POST['cantidad']),
                    observacion=request.POST.get('observacion', ''),
                    usuario=request.user
                )
                messages.success(request, 'Movimiento registrado exitosamente')
                return redirect('barra:movimientos_stock')
        except Exception as e:
            messages.error(request, f'Error: {e}')
    
    return redirect('barra:movimientos_stock')

# ===== CONTROL DE VENTAS =====

@barra_o_admin
def pos_ventas(request):
    """Punto de venta"""
    productos = Producto.objects.filter(activo=True, stock__gt=0).order_by('nombre')
    return render(request, 'core/barra/pos.html', {'productos': productos})

@require_http_methods(["POST"])
@ajax_login_required
@solo_barra
def procesar_venta(request):
    """Procesar venta en el POS"""
    try:
        data = json.loads(request.body)
        producto_id = data.get('producto_id')
        cantidad = int(data.get('cantidad', 1))
        asistente_cc = data.get('asistente_cc')
        usa_consumo = data.get('usa_consumo', False)
        
        if not producto_id or cantidad <= 0:
            return JsonResponse({'success': False, 'message': 'Datos inválidos'})
        
        try:
            producto = Producto.objects.get(pk=producto_id, activo=True)
        except Producto.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Producto no encontrado'})
        
        if producto.stock < cantidad:
            return JsonResponse({
                'success': False, 
                'message': f'Stock insuficiente. Disponible: {producto.stock}'
            })
        
        asistente = None
        if asistente_cc:
            try:
                asistente = Asistente.objects.get(cc=asistente_cc)
                if not asistente.ha_ingresado:
                    return JsonResponse({
                        'success': False,
                        'message': 'El asistente no ha ingresado al evento'
                    })
                
                if usa_consumo and asistente.consumos_disponibles < cantidad:
                    return JsonResponse({
                        'success': False,
                        'message': f'Consumos insuficientes. Disponibles: {asistente.consumos_disponibles}'
                    })
            except Asistente.DoesNotExist:
                return JsonResponse({'success': False, 'message': 'Asistente no encontrado'})
        
        # Procesar venta
        with transaction.atomic():
            venta = VentaBarra.objects.create(
                asistente=asistente,
                producto=producto,
                cantidad=cantidad,
                precio_unitario=producto.precio,
                usa_consumo_incluido=usa_consumo,
                vendedor=request.user
            )
        
        return JsonResponse({
            'success': True,
            'message': 'Venta procesada exitosamente',
            'venta': {
                'id': venta.pk,
                'total': float(venta.total),
                'producto': producto.nombre,
                'cantidad': cantidad,
                'stock_restante': producto.stock
            }
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'}, status=500)

@require_http_methods(["POST"])
@ajax_login_required
@solo_barra
def buscar_cliente_qr(request):
    """Buscar cliente por QR para ventas"""
    try:
        data = json.loads(request.body)
        codigo_qr = data.get('codigo_qr', '').strip()
        
        if not codigo_qr:
            return JsonResponse({'success': False, 'message': 'Código QR vacío'})
        
        try:
            asistente = Asistente.objects.get(codigo_qr=codigo_qr)
        except Asistente.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'QR no válido'})
        
        if not asistente.ha_ingresado:
            return JsonResponse({
                'success': False,
                'message': 'El asistente no ha ingresado al evento'
            })
        
        return JsonResponse({
            'success': True,
            'asistente': {
                'cc': asistente.cc,
                'nombre': asistente.nombre,
                'categoria': asistente.categoria.nombre,
                'consumos_disponibles': asistente.consumos_disponibles
            }
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'}, status=500)

# ===== REPORTES DE VENTAS =====

@barra_o_admin
def lista_ventas(request):
    """Lista de todas las ventas"""
    ventas = VentaBarra.objects.all().order_by('-fecha')
    
    # Filtrar por vendedor si es personal de barra
    try:
        if request.user.perfilusuario.rol == 'barra':
            ventas = ventas.filter(vendedor=request.user)
    except:
        pass
    
    # Filtros
    fecha = request.GET.get('fecha')
    if fecha:
        ventas = ventas.filter(fecha__date=fecha)
    
    producto_id = request.GET.get('producto')
    if producto_id:
        ventas = ventas.filter(producto_id=producto_id)
    
    context = {
        'ventas': ventas[:100],  # Limitar resultados
        'productos': Producto.objects.filter(activo=True),
        'filtros': {
            'fecha': fecha,
            'producto': producto_id
        }
    }
    
    return render(request, 'core/barra/lista_ventas.html', context)

@barra_o_admin
def ventas_del_dia(request):
    """Ventas del día actual"""
    hoy = datetime.now().date()
    ventas = VentaBarra.objects.filter(fecha__date=hoy)
    
    # Filtrar por vendedor si es personal de barra
    try:
        if request.user.perfilusuario.rol == 'barra':
            ventas = ventas.filter(vendedor=request.user)
    except:
        pass
    
    # Estadísticas
    stats = {
        'total_ventas': ventas.count(),
        'total_ingresos': ventas.aggregate(total=Sum('total'))['total'] or 0,
        'productos_vendidos': ventas.aggregate(total=Sum('cantidad'))['total'] or 0,
        'ventas_con_consumo': ventas.filter(usa_consumo_incluido=True).count(),
        'ventas_pagadas': ventas.filter(usa_consumo_incluido=False).count(),
    }
    
    # Productos más vendidos
    productos_vendidos = ventas.values('producto__nombre').annotate(
        total_vendido=Sum('cantidad'),
        ingresos=Sum('total')
    ).order_by('-total_vendido')
    
    context = {
        'ventas': ventas.order_by('-fecha'),
        'stats': stats,
        'productos_vendidos': productos_vendidos,
        'fecha': hoy
    }
    
    return render(request, 'core/barra/ventas_del_dia.html', context)

@barra_o_admin
def reportes_barra(request):
    """Reportes y estadísticas de barra"""
    hoy = datetime.now().date()
    
    # Estadísticas generales
    stats = {
        'productos_activos': Producto.objects.filter(activo=True).count(),
        'valor_inventario_actual': sum(p.precio * p.stock for p in Producto.objects.filter(activo=True)),
        'ventas_totales': VentaBarra.objects.count(),
        'ingresos_totales': VentaBarra.objects.aggregate(total=Sum('total'))['total'] or 0,
        'ventas_hoy': VentaBarra.objects.filter(fecha__date=hoy).count(),
        'ingresos_hoy': VentaBarra.objects.filter(fecha__date=hoy).aggregate(total=Sum('total'))['total'] or 0
    }
    
    # Top productos vendidos (todos los tiempos)
    top_productos = Producto.objects.annotate(
        total_vendido=Sum('ventabarra__cantidad'),
        ingresos_generados=Sum('ventabarra__total')
    ).filter(total_vendido__gt=0).order_by('-total_vendido')[:10]
    
    # Productos bajo stock
    productos_criticos = Producto.objects.filter(
        stock__lte=F('stock_minimo'), activo=True
    ).order_by('stock')
    
    context = {
        'stats': stats,
        'top_productos': top_productos,
        'productos_criticos': productos_criticos,
        'fecha_reporte': hoy
    }
    
    return render(request, 'core/barra/reportes.html', context)

@barra_o_admin
def reporte_final(request):
    """Reporte final del evento - inventario inicial vs final"""
    
    # Productos con stock inicial y actual
    productos_reporte = []
    for producto in Producto.objects.filter(activo=True):
        # Stock inicial (primer movimiento de entrada)
        primer_movimiento = MovimientoStock.objects.filter(
            producto=producto, tipo='entrada'
        ).order_by('fecha').first()
        
        stock_inicial = primer_movimiento.cantidad if primer_movimiento else 0
        
        # Total vendido
        total_vendido = VentaBarra.objects.filter(
            producto=producto
        ).aggregate(total=Sum('cantidad'))['total'] or 0
        
        # Ingresos generados
        ingresos = VentaBarra.objects.filter(
            producto=producto
        ).aggregate(total=Sum('total'))['total'] or 0
        
        productos_reporte.append({
            'producto': producto,
            'stock_inicial': stock_inicial,
            'stock_actual': producto.stock,
            'total_vendido': total_vendido,
            'ingresos_generados': ingresos,
            'diferencia_stock': stock_inicial - producto.stock - total_vendido
        })
    
    # Totales
    totales = {
        'valor_inventario_inicial': sum(p['stock_inicial'] * p['producto'].precio for p in productos_reporte),
        'valor_inventario_actual': sum(p['stock_actual'] * p['producto'].precio for p in productos_reporte),
        'total_vendido': sum(p['total_vendido'] for p in productos_reporte),
        'ingresos_totales': sum(p['ingresos_generados'] for p in productos_reporte)
    }
    
    context = {
        'productos_reporte': productos_reporte,
        'totales': totales,
        'fecha_reporte': datetime.now()
    }
    
    return render(request, 'core/barra/reporte_final.html', context)

# ===== EXPORTAR DATOS =====

@barra_o_admin
def exportar_inventario(request):
    """Exportar inventario actual a Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario DMT 69"
    
    # Headers
    headers = ['Producto', 'Precio', 'Stock Actual', 'Stock Mínimo', 'Valor Total', 'Estado']
    ws.append(headers)
    
    # Datos
    for producto in Producto.objects.filter(activo=True).order_by('nombre'):
        valor_total = producto.precio * producto.stock
        estado = 'OK' if producto.stock > producto.stock_minimo else 'BAJO STOCK'
        if producto.stock == 0:
            estado = 'AGOTADO'
            
        ws.append([
            producto.nombre,
            float(producto.precio),
            producto.stock,
            producto.stock_minimo,
            float(valor_total),
            estado
        ])
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=inventario_dmt69_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    wb.save(response)
    return response

@barra_o_admin
def exportar_ventas(request):
    """Exportar ventas a Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas DMT 69"
    
    # Headers
    headers = ['Fecha', 'Hora', 'Producto', 'Cantidad', 'Precio Unit.', 'Total', 'Cliente', 'Consumo Incluido', 'Vendedor']
    ws.append(headers)
    
    # Datos
    ventas = VentaBarra.objects.all().order_by('-fecha')
    for venta in ventas:
        ws.append([
            venta.fecha.strftime('%d/%m/%Y'),
            venta.fecha.strftime('%H:%M'),
            venta.producto.nombre,
            venta.cantidad,
            float(venta.precio_unitario),
            float(venta.total),
            venta.asistente.nombre if venta.asistente else 'Cliente General',
            'SÍ' if venta.usa_consumo_incluido else 'NO',
            venta.vendedor.username
        ])
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=ventas_dmt69_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    wb.save(response)
    return response

@barra_o_admin
def exportar_movimientos(request):
    """Exportar movimientos de stock a Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Movimientos Stock DMT 69"
    
    # Headers
    headers = ['Fecha', 'Producto', 'Tipo', 'Cantidad', 'Stock Anterior', 'Stock Nuevo', 'Observación', 'Usuario']
    ws.append(headers)
    
    # Datos
    movimientos = MovimientoStock.objects.all().order_by('-fecha')
    for mov in movimientos:
        ws.append([
            mov.fecha.strftime('%d/%m/%Y %H:%M'),
            mov.producto.nombre,
            mov.get_tipo_display(),
            mov.cantidad,
            mov.stock_anterior,
            mov.stock_nuevo,
            mov.observacion,
            mov.usuario.username
        ])
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=movimientos_dmt69_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    wb.save(response)
    return response