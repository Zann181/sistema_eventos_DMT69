import json
from datetime import datetime
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.db.models import Q, Count, Sum
from django.db import transaction
from .models import Asistente, Categoria, PerfilUsuario
from .decorators import entrada_o_admin, solo_entrada, ajax_login_required

from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone


from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
# Importar Decimal para convertir el precio pagado a un número entero
from decimal import Decimal, ROUND_HALF_UP

# Importar modelos desde core
from core.models import Producto, VentaBarra, MovimientoStock, Asistente, PerfilUsuario
from core.decorators import ajax_login_required


    # Agregar estas importaciones al inicio del archivo views.py
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
import io


# En core/views.py - Agregar esta función al final del archivo

from django.contrib.auth import logout

def custom_logout(request):
    """Vista personalizada de logout que garantiza redirección al login"""
    if request.user.is_authenticated:
        username = request.user.username
        logout(request)
        messages.success(request, f"Sesión de {username} cerrada exitosamente")
    return redirect('login')  # Garantiza que siempre vaya al login

# ===== DASHBOARD PRINCIPAL CORREGIDO =====
@login_required
def dashboard(request):
    """Dashboard principal que redirige según el rol del usuario"""
    try:
        perfil = request.user.perfilusuario
        if perfil.rol in ["entrada", "admin"]:
            return dashboard_entrada(request)
        elif perfil.rol == "barra":
            # TEMPORAL: usar dashboard_entrada hasta que esté listo el módulo barra
            messages.info(request, "🍺 Sistema de barra en desarrollo. Usando dashboard temporal.")
            return redirect('core:dashboard_barra')
        else:
            messages.error(request, "Rol no reconocido")
            return redirect("admin:index")
    except PerfilUsuario.DoesNotExist:
        messages.error(request, "Usuario sin perfil. Contacta al administrador.")
        return redirect("admin:index")





def dashboard_entrada(request):
    """
    Renderiza el dashboard de entrada con estadísticas y lista de categorías.

    Además de las estadísticas y la lista de categorías activas, esta vista
    prepara un diccionario con el precio de cada categoría para que el
    formulario de creación de asistentes pueda rellenar automáticamente el
    campo de «precio pagado» al seleccionar una categoría en el frontend.

    El diccionario `categoria_precios` se pasa en el contexto y se puede
    serializar a JSON dentro de la plantilla para utilizarlo en el script
    de autocompletado del precio.
    """
    hoy = datetime.now().date()

    # Estadísticas básicas
    stats = {
        "total_asistentes": Asistente.objects.count(),
        "asistentes_ingresados": Asistente.objects.filter(ha_ingresado=True).count(),
        "pendientes": Asistente.objects.filter(ha_ingresado=False).count(),
        "mis_verificaciones": Asistente.objects.filter(
            usuario_entrada=request.user, fecha_ingreso__date=hoy
        ).count(),
    }

    # Categorías activas con anotaciones de conteos básicos.
    # Luego se calcula manualmente el subtotal sumando el valor de consumo (consumos_disponibles)
    # de cada asistente en la categoría. Esto asegura que el subtotal refleje la suma correcta
    # sin depender de agregaciones condicionales que podrían fallar.
    categorias_qs = (
        Categoria.objects.filter(activa=True)
        .annotate(
            total=Count("asistente"),
            ingresados=Count("asistente", filter=Q(asistente__ha_ingresado=True)),
        )
    )

    categorias = []
    for c in categorias_qs:
        # Pendientes = total - ingresados
        c.pendientes = (c.total or 0) - (c.ingresados or 0)
        # Calcular subtotal sumando consumos_disponibles de todos los asistentes en la categoría
        suma = (
            Asistente.objects.filter(categoria=c)
            .aggregate(suma=Sum("consumos_disponibles"))
            .get("suma")
            or 0
        )
        c.subtotal = suma
        categorias.append(c)

    # Preparar diccionario de precios por categoría. Usamos str(c.pk) como clave
    # porque en la plantilla se tratará como cadena. Convertimos el precio a
    # cadena simple con str() para que no pierda decimales y sea seguro de
    # serializar. Si no hay precio definido, se asigna "0" por defecto.
    categoria_precios = {str(c.pk): str(c.precio) if c.precio is not None else "0" for c in categorias}

    # Serializar a JSON para que pueda ser insertado directamente en el
    # Javascript de la plantilla. Utilizamos el parámetro ensure_ascii=False
    # para permitir números con comas en locales que lo requieran y evitar
    # escapes innecesarios. En caso de error al serializar, se usará una
    # cadena vacía para evitar romper la vista.
    try:
        import json as _json  # local import para no contaminar espacio de nombres
        categoria_precios_json = _json.dumps(categoria_precios, ensure_ascii=False)
    except Exception:
        categoria_precios_json = '{}'

    # Total recaudado (suma de consumos disponibles de todos los asistentes)
    total_recaudado_data = Asistente.objects.aggregate(total=Sum("consumos_disponibles"))
    total_recaudado = total_recaudado_data.get("total") or 0

    # Número total de registros de asistentes
    registros = Asistente.objects.count()

    context = {
        "stats": stats,
        "categorias": categorias,
        "categoria_precios": categoria_precios,
        "categoria_precios_json": categoria_precios_json,
        "total_recaudado": total_recaudado,
        "registros": registros,
    }
    return render(request, "core/dashboard_entrada.html", context)


# ===== LISTA ASISTENTES =====
# En core/views.py - Reemplaza la función lista_asistentes existente
@entrada_o_admin
def lista_asistentes(request):
    asistentes = Asistente.objects.all().order_by("-fecha_registro")

    # Filtros
    buscar = request.GET.get("buscar", "").strip()
    if buscar:
        asistentes = asistentes.filter(
            Q(nombre__icontains=buscar) | Q(cc__icontains=buscar)
        )

    estado = request.GET.get("estado")
    if estado == "ingresados":
        asistentes = asistentes.filter(ha_ingresado=True)
    elif estado == "pendientes":
        asistentes = asistentes.filter(ha_ingresado=False)

    # Paginación
    items_por_pagina = int(request.GET.get('items', 10))  # Por defecto 10
    if items_por_pagina not in [10, 25, 50, 100]:
        items_por_pagina = 10
        
    paginator = Paginator(asistentes, items_por_pagina)
    page = request.GET.get('page', 1)
    
    try:
        asistentes_paginados = paginator.page(page)
    except PageNotAnInteger:
        asistentes_paginados = paginator.page(1)
    except EmptyPage:
        asistentes_paginados = paginator.page(paginator.num_pages)

    # Si es una petición AJAX, devolver solo la tabla
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        context = {
            "asistentes": asistentes_paginados,  # Ahora es paginado
            "paginator": paginator,
            "page_obj": asistentes_paginados,
            "total": paginator.count,
            "showing_start": asistentes_paginados.start_index(),
            "showing_end": asistentes_paginados.end_index(),
        }
        return render(request, "core/tabla_asistentes.html", context)

    context = {
        "asistentes": asistentes_paginados,
        "paginator": paginator,
        "page_obj": asistentes_paginados,
        "total": paginator.count,
        "buscar": buscar,
        "estado": estado,
        "items_por_pagina": items_por_pagina,
    }
    return render(request, "core/lista_asistentes.html", context)



# Reemplazar la función crear_asistente en views.py

@entrada_o_admin
def crear_asistente(request):
    if request.method == "POST":
        try:
            # Obtener el valor pagado enviado desde el formulario.  El campo se llama
            # "consumos_disponibles" en el formulario porque reutilizamos ese
            # atributo del modelo para almacenar el precio pagado.
            precio_str = request.POST.get("consumos_disponibles", "").strip()
            # Convertir a Decimal y luego a entero (pesos) redondeando al peso más cercano.
            # Si no se envía un valor o es inválido, se asignará 0 inicialmente y luego
            # se tomará el precio de la categoría como valor por defecto.
            consumo_disponible = 0
            if precio_str:
                try:
                    precio_decimal = Decimal(precio_str)
                    consumo_disponible = int(
                        precio_decimal.quantize(Decimal("1"), rounding=ROUND_HALF_UP)
                    )
                except Exception:
                    consumo_disponible = 0

            # Si no se proporcionó un precio válido, usar el precio de la categoría seleccionada
            if consumo_disponible == 0:
                categoria_id = request.POST.get("categoria")
                if categoria_id:
                    try:
                        categoria = Categoria.objects.get(pk=categoria_id)
                        # Convertir el precio de la categoría a entero redondeando al peso
                        precio_categoria = Decimal(categoria.precio or 0)
                        consumo_disponible = int(
                            precio_categoria.quantize(Decimal("1"), rounding=ROUND_HALF_UP)
                        )
                    except Exception:
                        # Si ocurre cualquier error (por ejemplo, categoría inexistente), dejar en 0
                        consumo_disponible = 0
            
            with transaction.atomic():
                cc = request.POST.get("cc", "").strip()
                if Asistente.objects.filter(cc=cc).exists():
                    messages.error(request, f"Ya existe asistente con cédula {cc}")
                else:
                    # Crear asistente, asignando consumos_disponibles al valor pagado
                    asistente = Asistente.objects.create(
                        nombre=request.POST["nombre"].strip(),
                        cc=cc,
                        numero=request.POST["numero"].strip(),
                        correo=request.POST["correo"].strip(),
                        categoria_id=request.POST["categoria"],
                        consumos_disponibles=consumo_disponible,
                        creado_por=request.user,
                    )
                    
                    # Generar QR si es necesario
                    if not asistente.qr_image:
                        asistente.generar_qr()
                    
                    # Enviar email de bienvenida al asistente
                    email_enviado = False
                    try:
                        from .email_utils import enviar_email_bienvenida
                        email_enviado = enviar_email_bienvenida(asistente)
                    except Exception as e:
                        # Imprimir el error en consola pero no interrumpir el flujo
                        print(f"Error enviando email: {e}")
                    
                    # Mostrar mensaje según resultado del envío de correo
                    if email_enviado:
                        messages.success(
                            request,
                            f"✅ Asistente {asistente.nombre} creado exitosamente. "
                            f"Email de confirmación enviado a {asistente.correo}"
                        )
                    else:
                        messages.warning(
                            request,
                            f"⚠️ Asistente {asistente.nombre} creado, pero no se pudo enviar el email. "
                            f"Verifica la configuración de correo en settings.py"
                        )
                    
                    return redirect("core:ver_qr", cc=asistente.cc)
        except Exception as e:
            # Capturar cualquier excepción y mostrar mensaje de error
            messages.error(request, f"Error al crear asistente: {e}")

    # Redirigir de vuelta al dashboard cuando no se envía un POST o se produce un error
    return redirect("core:dashboard")

# ===== EDITAR ASISTENTE =====
@entrada_o_admin
def editar_asistente(request, cc):
    asistente = get_object_or_404(Asistente, cc=cc)

    if request.method == "POST":
        try:
            with transaction.atomic():
                asistente.nombre = request.POST["nombre"].strip()
                asistente.numero = request.POST["numero"].strip()
                asistente.correo = request.POST["correo"].strip()
                asistente.categoria_id = request.POST["categoria"]
                # Actualizar consumos según nueva categoría
                asistente.consumos_disponibles = asistente.categoria.consumos_incluidos
                asistente.save()
                messages.success(request, f"Asistente {asistente.nombre} actualizado")
                return redirect("core:dashboard")
        except Exception as e:
            messages.error(request, f"Error: {e}")

    categorias = Categoria.objects.filter(activa=True)
    return render(
        request,
        "core/editar_asistente.html",
        {"asistente": asistente, "categorias": categorias},
    )


# ===== VER QR =====
@entrada_o_admin
def ver_qr(request, cc):
    asistente = get_object_or_404(Asistente, cc=cc)

    # Generar QR si no existe
    if not asistente.qr_image:
        asistente.generar_qr()

    return render(request, "core/ver_qr.html", {"asistente": asistente})


# ===== SCANNER QR =====
@solo_entrada
def scanner_qr(request):
    return render(request, "core/scanner.html")


@require_http_methods(["POST"])
@ajax_login_required
@solo_entrada
def verificar_qr(request):
    try:
        data = json.loads(request.body)
        codigo = data.get("codigo", "").strip()

        if not codigo:
            return JsonResponse({"success": False, "message": "Código QR vacío"})

        try:
            asistente = Asistente.objects.get(codigo_qr=codigo)
        except Asistente.DoesNotExist:
            return JsonResponse(
                {"success": False, "message": "Código QR inválido o no encontrado"}
            )

        # Verificar si ya ingresó
        if asistente.ha_ingresado:
            return JsonResponse(
                {
                    "success": False,
                    "message": f'{asistente.nombre} ya ingresó el {asistente.fecha_ingreso.strftime("%d/%m %H:%M")} verificado por {asistente.usuario_entrada.username}',
                }
            )

        # Marcar como ingresado
        with transaction.atomic():
            asistente.ha_ingresado = True
            asistente.fecha_ingreso = datetime.now()
            asistente.usuario_entrada = request.user
            asistente.save()

        return JsonResponse(
            {
                "success": True,
                "message": "ACCESO AUTORIZADO",
                "asistente": {
                    "nombre": asistente.nombre,
                    "cc": asistente.cc,
                    "categoria": asistente.categoria.nombre,
                    "consumos": asistente.consumos_disponibles,
                    "hora": datetime.now().strftime("%H:%M"),
                },
            }
        )

    except json.JSONDecodeError:
        return JsonResponse({"success": False, "message": "Datos JSON inválidos"})
    except Exception as e:
        return JsonResponse({"success": False, "message": f"Error interno: {str(e)}"})


# ===== BÚSQUEDA POR CC =====
@entrada_o_admin
def buscar_asistente(request):
    cc = request.GET.get("cc", "").strip()

    if not cc:
        return JsonResponse({"success": False, "message": "Número de cédula requerido"})

    try:
        asistente = Asistente.objects.get(cc=cc)
        return JsonResponse(
            {
                "success": True,
                "asistente": {
                    "nombre": asistente.nombre,
                    "cc": asistente.cc,
                    "categoria": asistente.categoria.nombre,
                    "ha_ingresado": asistente.ha_ingresado,
                    "fecha_ingreso": (
                        asistente.fecha_ingreso.strftime("%d/%m/%Y %H:%M")
                        if asistente.fecha_ingreso
                        else None
                    ),
                    "consumos": asistente.consumos_disponibles,
                    "verificado_por": (
                        asistente.usuario_entrada.username
                        if asistente.usuario_entrada
                        else None
                    ),
                },
            }
        )
    except Asistente.DoesNotExist:
        return JsonResponse(
            {"success": False, "message": "Asistente no encontrado con esa cédula"}
        )


# ===== MARCAR INGRESO MANUAL =====
@require_http_methods(["POST"])
@ajax_login_required
@entrada_o_admin
def marcar_ingreso(request):
    try:
        data = json.loads(request.body)
        cc = data.get("cc", "").strip()

        if not cc:
            return JsonResponse({"success": False, "message": "Cédula requerida"})

        try:
            asistente = Asistente.objects.get(cc=cc)
        except Asistente.DoesNotExist:
            return JsonResponse(
                {"success": False, "message": "Asistente no encontrado"}
            )

        if asistente.ha_ingresado:
            return JsonResponse(
                {"success": False, "message": "El asistente ya había ingresado"}
            )

        # Marcar como ingresado
        with transaction.atomic():
            asistente.ha_ingresado = True
            asistente.fecha_ingreso = datetime.now()
            asistente.usuario_entrada = request.user
            asistente.save()

        return JsonResponse(
            {
                "success": True,
                "message": f"{asistente.nombre} marcado como ingresado exitosamente",
            }
        )

    except Exception as e:
        return JsonResponse({"success": False, "message": f"Error: {str(e)}"})


# ===== ELIMINAR ASISTENTE =====
@require_http_methods(["POST"])
@ajax_login_required
@entrada_o_admin
def eliminar_asistente(request):
    try:
        data = json.loads(request.body)
        cc = data.get("cc", "").strip()

        if not cc:
            return JsonResponse({"success": False, "message": "Cédula requerida"})

        try:
            asistente = Asistente.objects.get(cc=cc)
        except Asistente.DoesNotExist:
            return JsonResponse(
                {"success": False, "message": "Asistente no encontrado"}
            )

        if asistente.ha_ingresado:
            return JsonResponse(
                {
                    "success": False,
                    "message": "No se puede eliminar un asistente que ya ingresó",
                }
            )

        nombre = asistente.nombre
        asistente.delete()

        return JsonResponse(
            {"success": True, "message": f"{nombre} eliminado exitosamente"}
        )

    except Exception as e:
        return JsonResponse({"success": False, "message": f"Error: {str(e)}"})


# ===== OBTENER QR =====
@entrada_o_admin
def obtener_qr(request, cc):
    try:
        asistente = Asistente.objects.get(cc=cc)
        if not asistente.qr_image:
            asistente.generar_qr()

        return JsonResponse(
            {
                "success": True,
                "qr_url": asistente.qr_image.url,
                "asistente": {
                    "nombre": asistente.nombre,
                    "cc": asistente.cc,
                    "categoria": asistente.categoria.nombre,
                    "consumos": asistente.consumos_disponibles,
                },
            }
        )
    except Asistente.DoesNotExist:
        return JsonResponse({"success": False, "message": "Asistente no encontrado"})


# VISTA PARA CONSULTAR QR (SIN GUARDAR NADA)
def buscar_asistente_qr(request):
    codigo_qr = request.GET.get("codigo_qr", None)
    if not codigo_qr:
        return JsonResponse({"success": False, "message": "Código QR no proporcionado"})

    try:
        asistente = Asistente.objects.get(codigo_qr=codigo_qr)

        if asistente.ha_ingresado:
            return JsonResponse(
                {
                    "success": False,
                    "message": f'{asistente.nombre} ya ingresó el {asistente.fecha_ingreso.strftime("%d/%m a las %H:%M")}',
                }
            )

        return JsonResponse(
            {
                "success": True,
                "asistente": {
                    "cc": asistente.cc,
                    "nombre": asistente.nombre,
                    "categoria": asistente.categoria.nombre,
                    "consumos_disponibles": asistente.consumos_disponibles,
                },
            }
        )
    except Asistente.DoesNotExist:
        return JsonResponse(
            {"success": False, "message": "Asistente no encontrado o QR inválido"}
        )


# En core/views.py
# ...


@require_http_methods(["POST"])
# ... (tus decoradores)
def verificar_qr(request):
    try:
        data = json.loads(request.body)
        # AHORA BUSCA POR CÉDULA (CC) EN LUGAR DE CÓDIGO QR
        cc = data.get("cc", "").strip()

        if not cc:
            return JsonResponse(
                {"success": False, "message": "Cédula no proporcionada"}
            )

        asistente = Asistente.objects.get(cc=cc)

        # ... (el resto de tu lógica para marcar el ingreso sigue igual) ...
        asistente.ha_ingresado = True
        asistente.fecha_ingreso = datetime.now()
        asistente.usuario_entrada = request.user
        asistente.save()

        return JsonResponse(
            {
                "success": True,
                "message": "Ingreso confirmado",
                "asistente": {"nombre": asistente.nombre},
            }
        )

    except Asistente.DoesNotExist:
        return JsonResponse(
            {"success": False, "message": "Asistente no encontrado para confirmar"}
        )
    except Exception as e:
        return JsonResponse({"success": False, "message": f"Error interno: {str(e)}"})




@require_http_methods(["POST"])
@ajax_login_required
@solo_entrada
def verificar_qr_preview(request):
    """Vista para obtener información del asistente sin marcar ingreso aún"""
    try:
        if request.content_type != 'application/json':
            return JsonResponse({'success': False, 'message': 'Tipo de contenido inválido'}, status=400)
        
        data = json.loads(request.body)
        codigo = data.get('codigo', '').strip()
        
        if not codigo:
            return JsonResponse({'success': False, 'message': 'Código QR vacío'})
        
        try:
            asistente = Asistente.objects.get(codigo_qr=codigo)
        except Asistente.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Código QR inválido o asistente no encontrado'})
        
        if asistente.ha_ingresado:
            return JsonResponse({
                'success': False,
                'message': 'Este asistente ya ingresó anteriormente',
                'asistente': {'nombre': asistente.nombre}
            })
        
        return JsonResponse({
            'success': True,
            'asistente': {
                'nombre': asistente.nombre,
                'cc': asistente.cc,
                'numero': asistente.numero,
                'correo': asistente.correo,
                'categoria': asistente.categoria.nombre,
                'consumos': asistente.consumos_disponibles,
                'fecha_registro': asistente.fecha_registro.strftime('%d/%m/%Y') if asistente.fecha_registro else None
            }
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'}, status=500)


@require_http_methods(["POST"])
@ajax_login_required
@solo_entrada
def confirmar_ingreso(request):
    """Vista para confirmar el ingreso después de la verificación"""
    try:
        if request.content_type != 'application/json':
            return JsonResponse({'success': False, 'message': 'Tipo de contenido inválido'}, status=400)
        
        data = json.loads(request.body)
        codigo = data.get('codigo', '').strip()
        
        if not codigo:
            return JsonResponse({'success': False, 'message': 'Código QR vacío'})
        
        try:
            asistente = Asistente.objects.get(codigo_qr=codigo)
        except Asistente.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Asistente no encontrado'})
        
        if asistente.ha_ingresado:
            return JsonResponse({'success': False, 'message': 'Ya ingresó anteriormente'})
        
        from django.utils import timezone
        fecha_ingreso = timezone.now()
        
        asistente.ha_ingresado = True
        asistente.fecha_ingreso = fecha_ingreso
        asistente.usuario_entrada = request.user
        asistente.save()
        
        return JsonResponse({
            'success': True,
            'asistente': {
                'nombre': asistente.nombre,
                'cc': asistente.cc,
                'categoria': asistente.categoria.nombre,
                'consumos': asistente.consumos_disponibles,
                'hora': fecha_ingreso.strftime('%H:%M:%S'),
                'fecha': fecha_ingreso.strftime('%d/%m/%Y')
            }
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'}, status=500)
    

# Corrige la vista exportar_excel en views.py
@entrada_o_admin
def exportar_excel(request):
    """Exportar todos los asistentes a un archivo Excel"""
    try:
        # Obtener todos los asistentes
        asistentes = Asistente.objects.all().order_by('-fecha_registro')
        
        # Crear workbook y worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Asistentes DMT 69"
        
        # Configurar estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        # Definir headers
        headers = [
            'NOMBRE', 'CEDULA', 'TELEFONO', 'CORREO', 'CATEGORIA', 
            'CONSUMOS DISPONIBLES', 'ESTADO', 'FECHA REGISTRO', 
            'FECHA INGRESO', 'HORA INGRESO', 'VERIFICADO POR'
        ]
        
        # Escribir headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        
        # Escribir datos de asistentes
        for row_num, asistente in enumerate(asistentes, 2):
            ws.cell(row=row_num, column=1, value=asistente.nombre)
            ws.cell(row=row_num, column=2, value=asistente.cc)
            ws.cell(row=row_num, column=3, value=asistente.numero or '')
            ws.cell(row=row_num, column=4, value=asistente.correo or '')
            ws.cell(row=row_num, column=5, value=asistente.categoria.nombre)
            ws.cell(row=row_num, column=6, value=asistente.consumos_disponibles)
            ws.cell(row=row_num, column=7, value="INGRESO" if asistente.ha_ingresado else "PENDIENTE")  # Corregido el texto
            
            # Fecha de registro
            if asistente.fecha_registro:
                ws.cell(row=row_num, column=8, value=asistente.fecha_registro.strftime('%d/%m/%Y'))
            else:
                ws.cell(row=row_num, column=8, value='')
            
            # Fecha y hora de ingreso
            if asistente.fecha_ingreso:
                ws.cell(row=row_num, column=9, value=asistente.fecha_ingreso.strftime('%d/%m/%Y'))
                ws.cell(row=row_num, column=10, value=asistente.fecha_ingreso.strftime('%H:%M:%S'))
            else:
                ws.cell(row=row_num, column=9, value='')
                ws.cell(row=row_num, column=10, value='')
            
            # Usuario que verificó
            if asistente.usuario_entrada:
                ws.cell(row=row_num, column=11, value=asistente.usuario_entrada.username)
            else:
                ws.cell(row=row_num, column=11, value='')
        
        # Ajustar ancho de columnas
        column_widths = {
            'A': 25,  # NOMBRE
            'B': 15,  # CEDULA
            'C': 15,  # TELEFONO
            'D': 30,  # CORREO
            'E': 20,  # CATEGORIA
            'F': 12,  # CONSUMOS
            'G': 12,  # ESTADO
            'H': 15,  # FECHA REGISTRO
            'I': 15,  # FECHA INGRESO
            'J': 12,  # HORA INGRESO
            'K': 15,  # VERIFICADO POR
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Agregar información adicional al final
        total_row = len(asistentes) + 3
        ws.cell(row=total_row, column=1, value="RESUMEN:")
        ws.cell(row=total_row, column=1).font = Font(bold=True)
        
        ws.cell(row=total_row + 1, column=1, value=f"Total asistentes: {asistentes.count()}")
        ws.cell(row=total_row + 2, column=1, value=f"Han ingresado: {asistentes.filter(ha_ingresado=True).count()}")
        ws.cell(row=total_row + 3, column=1, value=f"Pendientes: {asistentes.filter(ha_ingresado=False).count()}")
        ws.cell(row=total_row + 4, column=1, value=f"Exportado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
        ws.cell(row=total_row + 5, column=1, value=f"Exportado por: {request.user.username}")
        
        # Crear respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        filename = f"asistentes_dmt69_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Guardar workbook en response
        wb.save(response)
        
        return response
        
    except Exception as e:
        # En caso de error, devolver respuesta de error
        return JsonResponse({
            'success': False, 
            'message': f'Error al generar Excel: {str(e)}'
        }, status=500)

@entrada_o_admin
def exportar_excel_simple(request):
    """Versión simple para testing"""
    import csv
    from django.http import HttpResponse
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="asistentes.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Nombre', 'Cedula', 'Telefono', 'Email', 'Categoria', 'Estado'])
    
    asistentes = Asistente.objects.all()
    for asistente in asistentes:
        writer.writerow([
            asistente.nombre,
            asistente.cc,
            asistente.numero,
            asistente.correo,
            asistente.categoria.nombre,
            'INGRESÓ' if asistente.ha_ingresado else 'PENDIENTE'
        ])
    
    return response




# --------- Barras y ventas -------
# BUSCA la función dashboard_barra en core/views.py (línea ~655) y REEMPLÁZALA con esta:

@login_required
def dashboard_barra(request):
    """Dashboard del sistema de barra con cálculo basado en descripción"""
    try:
        perfil = request.user.perfilusuario
        if perfil.rol not in ['barra', 'admin']:
            messages.error(request, 'No tienes permisos para acceder al sistema de barra')
            return redirect('core:dashboard')
    except PerfilUsuario.DoesNotExist:
        messages.error(request, 'Usuario sin perfil asignado')
        return redirect('admin:index')
    
    # Calcular estadísticas basadas en la diferencia de stock
    stats = calcular_stats_ventas_hoy(request.user)
    
    # Lista de productos con información de stock
    productos_info = []
    try:
        productos = Producto.objects.filter(activo=True).order_by('nombre')
        
        for producto in productos:
            # Determinar color del stock
            if producto.stock == 0:
                color_stock = 'danger'
            elif producto.stock <= producto.stock_minimo:
                color_stock = 'warning'
            else:
                color_stock = 'success'
            
            # Calcular productos vendidos para este producto específico
            stock_inicial = extraer_stock_inicial_de_descripcion(producto.descripcion)
            productos_vendidos_hoy = max(0, stock_inicial - producto.stock) if stock_inicial is not None else 0
            
            productos_info.append({
                'producto': producto,
                'color_stock': color_stock,
                'puede_vender': producto.stock > 0,
                'stock_inicial_dia': stock_inicial,
                'vendidos_hoy': productos_vendidos_hoy
            })
    except Exception as e:
        print(f"Error cargando productos: {e}")
        productos_info = []
    
    context = {
        'stats': stats,
        'productos': productos_info,
        'user': request.user
    }
    
    return render(request, 'core/dashboard_barra.html', context)

import re


def extraer_stock_inicial_de_descripcion(descripcion: str):
    """
    Devuelve el total de la descripción sumando todos los números que encuentre.
    Ej.: "300+50+120" -> 470. Si no hay números, retorna None.
    """
    if not descripcion:
        return None
    numeros = re.findall(r"\d+", str(descripcion))
    if not numeros:
        return None
    return sum(int(n) for n in numeros)



def calcular_stats_ventas_hoy(usuario):
    """
    Calcula las estadísticas de ventas del día basándose en la diferencia de stock
    """
    try:
        productos = Producto.objects.filter(activo=True)
        
        total_productos_vendidos = 0
        total_ingresos_estimados = 0.0
        productos_con_ventas = 0
        
        for producto in productos:
            # Extraer stock inicial de la descripción
            stock_inicial = extraer_stock_inicial_de_descripcion(producto.descripcion)
            
            if stock_inicial is not None:
                # Calcular productos vendidos (diferencia entre inicial y actual)
                vendidos_hoy = max(0, stock_inicial - producto.stock)
                
                if vendidos_hoy > 0:
                    total_productos_vendidos += vendidos_hoy
                    total_ingresos_estimados += vendidos_hoy * float(producto.precio)
                    productos_con_ventas += 1
        
        # Contar productos activos y agotados
        productos_activos = productos.filter(stock__gt=0).count()
        productos_agotados = productos.filter(stock=0).count()
        
        return {
            'productos_activos': productos_activos,
            'productos_agotados': productos_agotados,
            'mis_ventas_hoy': productos_con_ventas,  # Productos diferentes vendidos
            'productos_vendidos_hoy': total_productos_vendidos,  # Cantidad total vendida
            'total_vendido_hoy': round(total_ingresos_estimados, 2)
        }
        
    except Exception as e:
        print(f"Error calculando stats de ventas: {e}")
        return {
            'productos_activos': 0,
            'productos_agotados': 0,
            'mis_ventas_hoy': 0,
            'productos_vendidos_hoy': 0,
            'total_vendido_hoy': 0.0
        }


@ajax_login_required
def obtener_stats_barra(request):
    """Obtener estadísticas actualizadas para el dashboard"""
    try:
        stats = calcular_stats_ventas_hoy(request.user)
        return JsonResponse({'success': True, 'stats': stats})
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})


@login_required
def mis_ventas_barra(request):
    """Ver mis ventas del día calculadas por diferencia de stock"""
    try:
        perfil = request.user.perfilusuario
        if perfil.rol not in ['barra', 'admin']:
            messages.error(request, 'No tienes permisos para ver ventas')
            return redirect('core:dashboard')
    except PerfilUsuario.DoesNotExist:
        messages.error(request, 'Usuario sin perfil asignado')
        return redirect('admin:index')
    
    fecha_filtro = datetime.now().date()
    
    # Calcular ventas basándose en la diferencia de stock
    try:
        productos = Producto.objects.filter(activo=True)
        ventas_calculadas = []
        
        total_productos_vendidos = 0
        total_ingresos = 0.0
        total_transacciones = 0
        
        for producto in productos:
            stock_inicial = extraer_stock_inicial_de_descripcion(producto.descripcion)
            
            if stock_inicial is not None:
                vendidos_hoy = max(0, stock_inicial - producto.stock)
                
                if vendidos_hoy > 0:
                    ingreso_producto = vendidos_hoy * float(producto.precio)
                    
                    # Crear objeto simulado de venta para mostrar en la template
                    venta_simulada = {
                        'producto': producto,
                        'cantidad': vendidos_hoy,
                        'precio_unitario': producto.precio,
                        'total': ingreso_producto,
                        'fecha': datetime.now(),  # Fecha actual como aproximación
                        'asistente': None,  # Cliente general
                        'usa_consumo_incluido': False
                    }
                    
                    ventas_calculadas.append(venta_simulada)
                    total_productos_vendidos += vendidos_hoy
                    total_ingresos += ingreso_producto
                    total_transacciones += 1
        
        stats_ventas = {
            'total_ventas': total_transacciones,
            'productos_vendidos': total_productos_vendidos,
            'total_ingresos': round(total_ingresos, 2)
        }
        
    except Exception as e:
        print(f"Error calculando ventas: {e}")
        ventas_calculadas = []
        stats_ventas = {
            'total_ventas': 0,
            'productos_vendidos': 0,
            'total_ingresos': 0
        }
    
    context = {
        'ventas': ventas_calculadas,
        'stats_ventas': stats_ventas,
        'fecha': fecha_filtro,
        'user': request.user,
        'es_calculo_estimado': True  # Flag para mostrar en template que son cálculos
    }
    
    return render(request, 'core/mis_ventas_barra.html', context)


# FUNCIÓN AUXILIAR para actualizar stock inicial en descripción (opcional)
@login_required
def actualizar_stock_inicial_dia(request):
    """
    Función auxiliar para actualizar el stock inicial del día en las descripciones
    Llamar al inicio del día para resetear los contadores
    """
    try:
        perfil = request.user.perfilusuario
        if perfil.rol not in ['admin']:
            return JsonResponse({'success': False, 'message': 'Solo administradores'})
    except PerfilUsuario.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Usuario sin perfil'})
    
    try:
        productos_actualizados = 0
        
        for producto in Producto.objects.filter(activo=True):
            # Actualizar la descripción con el stock actual como inicial del día
            nueva_descripcion = f"Stock inicial día: {producto.stock}"
            
            # Mantener descripción anterior si existe y no contiene info de stock
            if producto.descripcion and 'stock inicial' not in producto.descripcion.lower():
                nueva_descripcion = f"{producto.descripcion} | {nueva_descripcion}"
            
            producto.descripcion = nueva_descripcion
            producto.save(update_fields=['descripcion'])
            productos_actualizados += 1
        
        return JsonResponse({
            'success': True, 
            'message': f'Stock inicial actualizado para {productos_actualizados} productos'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})


# FUNCIÓN DE DEBUG para verificar cálculos
@login_required
def debug_ventas_calculadas(request):
    """Función de debug para verificar los cálculos de ventas"""
    try:
        productos_debug = []
        
        for producto in Producto.objects.filter(activo=True):
            stock_inicial = extraer_stock_inicial_de_descripcion(producto.descripcion)
            vendidos = max(0, stock_inicial - producto.stock) if stock_inicial else 0
            
            productos_debug.append({
                'nombre': producto.nombre,
                'descripcion': producto.descripcion,
                'stock_actual': producto.stock,
                'stock_inicial_extraido': stock_inicial,
                'vendidos_calculados': vendidos,
                'ingresos_calculados': vendidos * float(producto.precio) if vendidos > 0 else 0
            })
        
        return JsonResponse({
            'productos': productos_debug,
            'timestamp': datetime.now().strftime('%H:%M:%S')
        })
        
    except Exception as e:
        return JsonResponse({'error': f'Error en debug: {str(e)}'})







# REEMPLAZA LA FUNCIÓN vender_producto EN core/views.py

from django.db import transaction

@require_http_methods(["POST"])
@login_required
def vender_producto(request):
    """Procesa la venta de 1 unidad, descuenta stock una sola vez y registra la venta."""
    # Leer JSON
    try:
        data = json.loads(request.body or b"{}")
        producto_id = int(data.get('producto_id', 0))
        cantidad = int(data.get('cantidad', 1) or 1)
    except Exception:
        return JsonResponse({'success': False, 'message': 'Payload inválido'}, status=400)

    if producto_id <= 0 or cantidad <= 0:
        return JsonResponse({'success': False, 'message': 'Producto o cantidad inválidos'}, status=400)

    try:
        with transaction.atomic():
            producto = Producto.objects.select_for_update().get(id=producto_id, activo=True)
            if producto.stock < cantidad:
                return JsonResponse({'success': False, 'message': f'Stock insuficiente. Disponible: {producto.stock}'})

            # Descontar exactamente una vez
            producto.stock -= cantidad
            producto.save(update_fields=['stock'])

            venta = VentaBarra.objects.create(
                asistente=None,               # ajusta si corresponde
                producto=producto,
                cantidad=cantidad,
                precio_unitario=producto.precio,
                total=producto.precio * cantidad,
                vendedor=request.user,
            )

        # Color del stock para UI
        if producto.stock == 0:
            color = 'danger'
        elif producto.stock <= getattr(producto, 'stock_minimo', 0):
            color = 'warning'
        else:
            color = 'success'

        return JsonResponse({
            'success': True,
            'message': f'Venta de {producto.nombre} x{cantidad} exitosa',
            'venta': {
                'id': venta.id,
                'producto_nombre': producto.nombre,
                'cantidad': cantidad,
                'total': float(venta.total),
                'nuevo_stock': producto.stock,
                'color_stock': color,
                'puede_vender': producto.stock > 0,
                'hora': getattr(venta, 'fecha', None).strftime('%H:%M:%S') if getattr(venta, 'fecha', None) else '',
            },
        })

    except Producto.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Producto no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error: {e}'}, status=500)


# OPCIONAL: Función de DEBUG para verificar stock después de ventas
@login_required
def debug_stock(request, producto_id):
    """Función temporal para verificar el stock de un producto"""
    try:
        producto = Producto.objects.get(id=producto_id)
        return JsonResponse({
            'producto': producto.nombre,
            'stock_actual': producto.stock,
            'timestamp': datetime.now().strftime('%H:%M:%S')
        })
    except Producto.DoesNotExist:
        return JsonResponse({'error': 'Producto no encontrado'})


@ajax_login_required
def buscar_asistente_barra(request):
    """Buscar asistente por cédula para la venta"""
    cc = request.GET.get('cc', '').strip()
    
    if not cc:
        return JsonResponse({'success': False, 'message': 'Cédula requerida'})
    
    try:
        asistente = Asistente.objects.get(cc=cc)
        
        if not asistente.ha_ingresado:
            return JsonResponse({
                'success': False,
                'message': f'{asistente.nombre} no ha ingresado al evento'
            })
        
        return JsonResponse({
            'success': True,
            'asistente': {
                'nombre': asistente.nombre,
                'cc': asistente.cc,
                'categoria': asistente.categoria.nombre,
                'consumos_disponibles': asistente.consumos_disponibles,
                'puede_usar_consumo': asistente.consumos_disponibles > 0
            }
        })
        
    except Asistente.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Asistente no encontrado'})




@login_required
def exportar_reporte_barra(request):
    """Exportar reporte de ventas a Excel basado en cálculo de diferencias de stock"""
    try:
        perfil = request.user.perfilusuario
        if perfil.rol not in ['barra', 'admin']:
            messages.error(request, 'No tienes permisos')
            return redirect('core:dashboard')
    except PerfilUsuario.DoesNotExist:
        messages.error(request, 'Usuario sin perfil')
        return redirect('admin:index')
    
    try:
        hoy = datetime.now().date()
        
        # Calcular ventas basándose en la diferencia de stock (igual que en mis_ventas_barra)
        productos = Producto.objects.filter(activo=True)
        ventas_calculadas = []
        
        total_productos_vendidos = 0
        total_ingresos = 0.0
        total_transacciones = 0
        
        for producto in productos:
            stock_inicial = extraer_stock_inicial_de_descripcion(producto.descripcion)
            
            if stock_inicial is not None:
                vendidos_hoy = max(0, stock_inicial - producto.stock)
                
                if vendidos_hoy > 0:
                    ingreso_producto = vendidos_hoy * float(producto.precio)
                    
                    # Crear objeto simulado de venta para el Excel
                    venta_simulada = {
                        'producto_nombre': producto.nombre,
                        'cantidad': vendidos_hoy,
                        'precio_unitario': float(producto.precio),
                        'total': ingreso_producto,
                        'stock_inicial': stock_inicial,
                        'stock_actual': producto.stock,
                        'fecha': datetime.now()  # Fecha actual como aproximación
                    }
                    
                    ventas_calculadas.append(venta_simulada)
                    total_productos_vendidos += vendidos_hoy
                    total_ingresos += ingreso_producto
                    total_transacciones += 1
        
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Ventas {request.user.username}"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        # Headers actualizados para incluir información de stock
        headers = ['PRODUCTO', 'STOCK INICIAL', 'STOCK ACTUAL', 'CANTIDAD VENDIDA', 'PRECIO UNIT.', 'TOTAL INGRESOS', 'METODO']
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        
        # Datos calculados
        for row_num, venta in enumerate(ventas_calculadas, 2):
            ws.cell(row=row_num, column=1, value=venta['producto_nombre'])
            ws.cell(row=row_num, column=2, value=venta['stock_inicial'])
            ws.cell(row=row_num, column=3, value=venta['stock_actual'])
            ws.cell(row=row_num, column=4, value=venta['cantidad'])
            ws.cell(row=row_num, column=5, value=venta['precio_unitario'])
            ws.cell(row=row_num, column=6, value=venta['total'])
            ws.cell(row=row_num, column=7, value="Cálculo automático")
        
        # Ajustar columnas
        column_widths = [25, 15, 15, 15, 15, 18, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Totales
        total_row = len(ventas_calculadas) + 3
        ws.cell(row=total_row, column=1, value="TOTALES:")
        ws.cell(row=total_row, column=1).font = Font(bold=True)
        
        ws.cell(row=total_row, column=4, value=total_productos_vendidos)
        ws.cell(row=total_row, column=6, value=total_ingresos)
        
        # Información adicional
        ws.cell(row=total_row + 2, column=1, value=f"Fecha: {hoy.strftime('%d/%m/%Y')}")
        ws.cell(row=total_row + 3, column=1, value=f"Vendedor: {request.user.username}")
        ws.cell(row=total_row + 4, column=1, value=f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
        ws.cell(row=total_row + 5, column=1, value="Método: Cálculo basado en diferencia de stock")
        
        # Resumen de estadísticas
        ws.cell(row=total_row + 7, column=1, value="RESUMEN:")
        ws.cell(row=total_row + 7, column=1).font = Font(bold=True)
        ws.cell(row=total_row + 8, column=1, value=f"Productos diferentes vendidos: {total_transacciones}")
        ws.cell(row=total_row + 9, column=1, value=f"Unidades totales vendidas: {total_productos_vendidos}")
        ws.cell(row=total_row + 10, column=1, value=f"Ingresos totales estimados: ${total_ingresos:.2f}")
        
        # Nota explicativa
        ws.cell(row=total_row + 12, column=1, value="NOTA:")
        ws.cell(row=total_row + 12, column=1).font = Font(bold=True)
        ws.cell(row=total_row + 13, column=1, value="Las ventas se calculan automáticamente comparando")
        ws.cell(row=total_row + 14, column=1, value="el stock inicial del día (registrado en descripción)")
        ws.cell(row=total_row + 15, column=1, value="con el stock actual de cada producto.")
        
        # Respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"ventas_calculadas_{request.user.username}_{hoy.strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        messages.error(request, f'Error generando reporte: {str(e)}')
        return redirect('core:mis_ventas_barra')    



def _check_rol_barra_o_admin(user):
    try:
        return user.perfilusuario.rol in ['barra', 'admin']
    except:
        return False

@require_http_methods(["POST"])
@login_required
def crear_producto_rapido(request):
    """Crear producto sencillo desde la pestaña (solo nombre, precio, stock, stock_minimo, descripcion)."""
    if not _check_rol_barra_o_admin(request.user):
        return JsonResponse({'success': False, 'message': 'Sin permisos'}, status=403)

    try:
        nombre = request.POST.get('nombre', '').strip()
        precio = float(request.POST.get('precio', 0))
        stock = int(request.POST.get('stock', 0))
        stock_minimo = int(request.POST.get('stock_minimo', 5))
        descripcion = request.POST.get('descripcion', '').strip()

        if not nombre or precio < 0 or stock < 0 or stock_minimo < 0:
            raise ValueError("Datos inválidos")

        with transaction.atomic():
            p = Producto.objects.create(
                nombre=nombre,
                precio=precio,
                stock=stock,
                stock_minimo=stock_minimo,
                descripcion=descripcion,
                creado_por=request.user,
                activo=True,
            )
        # vuelve al dashboard de barra
        from django.contrib import messages
        messages.success(request, f"Producto {p.nombre} creado correctamente")
        return redirect('core:dashboard_barra')
    except Exception as e:
        from django.contrib import messages
        messages.error(request, f"Error: {e}")
        return redirect('core:dashboard_barra')



from django.views.decorators.http import require_http_methods
from django.contrib import messages
@require_http_methods(["POST"])
@login_required
def sumar_descripcion_producto(request):
    """
    Agrega "+cantidad" a la descripción y suma esa cantidad al stock real del producto.
    Acepta tanto JSON (AJAX) como Form-POST (desde el modal).
    """
    def wants_json(req):
        ct = (req.content_type or "")
        return req.headers.get("x-requested-with") == "XMLHttpRequest" or ct.startswith("application/json")

    def respond(ok, status=200, **payload):
        if wants_json(request):
            return JsonResponse({"success": ok, **payload}, status=status)
        # POST normal: redirigir con mensajes flash
        if ok:
            messages.success(request, f"Descripción actualizada a '{payload.get('nueva_descripcion','')}'. Stock: {payload.get('nuevo_stock','?')}")
        else:
            messages.error(request, payload.get('message', 'Error al actualizar'))
        return redirect('core:dashboard_barra')

    try:
        # Extraer datos (JSON o Form)
        if (request.content_type or "").startswith("application/json"):
            data = json.loads((request.body or b"{}").decode("utf-8"))
            producto_id = int(data.get("producto_id", 0))
            cantidad = int(data.get("cantidad", 0))
        else:
            producto_id = int(request.POST.get("producto_id", 0))
            cantidad = int(request.POST.get("cantidad", 0))

        if producto_id <= 0 or cantidad <= 0:
            return respond(False, status=400, message="Producto o cantidad inválidos")

        with transaction.atomic():
            # Lock del producto
            p = Producto.objects.select_for_update().get(pk=producto_id, activo=True)

            # Construir la nueva descripción
            desc_actual = (p.descripcion or "").strip()
            nueva_desc = str(cantidad) if not desc_actual else f"{desc_actual}+{cantidad}"

            # Incrementar STOCK real
            p.descripcion = nueva_desc
            p.stock = (p.stock or 0) + cantidad

            # Guardar cambios
            try:
                p.save(update_fields=["descripcion", "stock"])  # añade aquí otros campos si tu modelo los usa (p.ej. fecha_actualizacion)
            except Exception:
                p.save()

        # Color para la tarjeta
        if p.stock == 0:
            color = "danger"
        elif p.stock <= getattr(p, "stock_minimo", 0):
            color = "warning"
        else:
            color = "success"

        return respond(True,
                       nueva_descripcion=nueva_desc,
                       nuevo_stock=p.stock,
                       color_stock=color,
                       puede_vender=p.stock > 0)

    except Producto.DoesNotExist:
        return respond(False, status=404, message="Producto no encontrado")
    except json.JSONDecodeError:
        return respond(False, status=400, message="JSON inválido")
    except Exception as e:
        return respond(False, status=400, message=f"Error: {e}")