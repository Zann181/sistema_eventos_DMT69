import json
from decimal import Decimal
from urllib.parse import quote

import openpyxl
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.db.models import Count, Q, Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_GET, require_POST
from openpyxl.styles import Alignment, Font, PatternFill
from attendees.application import check_in_attendee, delete_branch_category, get_attendee_for_branch
from attendees.forms import AttendeeForm, BranchCategoryForm
from attendees.models import Attendee, Category
from identity.application import user_can_access_attendees, user_can_manage_categories, user_can_manage_events
from sales.application import (
    create_cash_movement,
    delete_cash_movement,
    extract_split_payments,
    register_event_day_entry,
    resolve_expense_payments,
    update_cash_movement,
)
from sales.forms import CashDropForm, EventDayEntryForm, ExpenseForm
from sales.models import CashMovement, CashMovementPayment
from ticketing.application import (
    build_event_share_text,
    build_qr_png_bytes,
    build_whatsapp_share_card_png,
    send_attendee_ticket_email,
)
from media_assets.application import resolve_field_file


ATTENDEES_CONTENT_TABS = {"scanner", "lista", "crear"}
ATTENDEES_MODAL_TABS = {"categorias", "evento-dia", "gastos", "vaciar-caja"}
ATTENDEES_RETURN_TABS = ATTENDEES_CONTENT_TABS | ATTENDEES_MODAL_TABS


def _sanitize_attendees_content_tab(value, default="scanner"):
    return value if value in ATTENDEES_CONTENT_TABS else default


def _sanitize_attendees_modal_tab(value):
    return value if value in ATTENDEES_MODAL_TABS else ""


def _sanitize_attendees_return_tab(value, default="scanner"):
    return value if value in ATTENDEES_RETURN_TABS else default


def _build_attendee_form(branch, event, data=None, selected_category_id=None):
    initial = {}
    if data is None and selected_category_id:
        category = branch.categories.filter(pk=selected_category_id, is_active=True).first()
        if category:
            initial["category"] = category.pk
            initial["paid_amount"] = category.price
    return AttendeeForm(data or None, branch=branch, event=event, initial=initial)


def _get_editing_category(branch, category_id):
    if not category_id:
        return None
    return branch.categories.filter(pk=category_id).first()


def _get_branch_and_event(request):
    branch = request.current_branch
    event = request.current_event
    if not branch or not event:
        messages.error(request, "Selecciona sucursal y evento.")
        return None, None
    return branch, event


def _ensure_attendee_access(request, branch, event):
    if user_can_access_attendees(request.user, branch, event):
        return True
    messages.error(request, "No tienes permisos para acceder al modulo de entrada.")
    return False


def _ensure_category_management_access(request, branch, event):
    if user_can_manage_categories(request.user, branch, event):
        return True
    messages.error(request, "Solo los administradores pueden gestionar categorias.")
    return False


def _ensure_cash_management_access(request, branch, event):
    if user_can_manage_events(request.user, branch, event):
        return True
    messages.error(request, "Solo los administradores pueden gestionar gastos y vaciados.")
    return False


def _get_editing_cash_movement(branch, event, module, movement_type, movement_id):
    if not movement_id:
        return None
    return (
        CashMovement.objects.filter(
            pk=movement_id,
            branch=branch,
            event=event,
            module=module,
            movement_type=movement_type,
        )
        .select_related("created_by")
        .prefetch_related("payments")
        .first()
    )


def _build_expense_form(movement=None):
    initial = {}
    if movement is not None:
        initial = {
            "amount": movement.total_amount,
            "description": movement.description,
        }
    return ExpenseForm(initial=initial)


def _build_cash_drop_form(movement=None):
    initial = {}
    if movement is not None:
        initial = {
            "amount": movement.total_amount,
            "description": movement.description,
        }
    return CashDropForm(initial=initial)


def _expense_payment_inputs_present(request, *, prefix):
    if (request.POST.get("payment_method") or "").strip():
        return True
    for index in range(1, 5):
        if (
            (request.POST.get(f"{prefix}_payment_method_{index}") or "").strip()
            or (request.POST.get(f"{prefix}_payment_amount_{index}") or "").strip()
            or (request.POST.get(f"{prefix}_payment_reference_{index}") or "").strip()
            or request.FILES.get(f"{prefix}_payment_proof_{index}")
        ):
            return True
    return False


def _attendee_queryset(branch, event):
    return (
        Attendee.objects.filter(branch=branch, event=event)
        .select_related("category", "checked_in_by")
        .order_by("-created_at")
    )


def _normalize_whatsapp_phone(raw_phone):
    digits = "".join(char for char in str(raw_phone or "") if char.isdigit())
    if digits.startswith("00"):
        digits = digits[2:]
    if not digits:
        return ""
    if not digits.startswith("57"):
        digits = f"57{digits.lstrip('0')}"
    return digits


def _format_whatsapp_phone_display(raw_phone):
    normalized = _normalize_whatsapp_phone(raw_phone)
    if len(normalized) >= 12 and normalized.startswith("57"):
        national_number = normalized[2:]
        if len(national_number) >= 10:
            return f"+57 {national_number[:3]} {national_number[3:6]} {national_number[6:]}"
        return f"+57 {national_number}"
    if normalized:
        return f"+{normalized}"
    return ""


def _build_media_url(request, field_file):
    if not field_file:
        return ""
    try:
        relative_url = field_file.url
    except ValueError:
        return ""
    public_base_url = (
        getattr(settings, "WHATSAPP_MEDIA_BASE_URL", "")
        or getattr(settings, "EMAIL_MEDIA_BASE_URL", "")
    ).rstrip("/")
    if public_base_url:
        return f"{public_base_url}{relative_url}"
    return request.build_absolute_uri(relative_url)


def _build_whatsapp_url(request, attendee):
    phone = _normalize_whatsapp_phone(attendee.phone)
    qr_url = _build_whatsapp_file_url(request, "attendees:whatsapp_qr_file", attendee.qr_code)
    text = quote(build_event_share_text(attendee.event, attendee, qr_url=qr_url))

    if phone:
        return f"https://web.whatsapp.com/send?phone={phone}&text={text}"
    return f"https://web.whatsapp.com/send?text={text}"


def _build_public_absolute_url(request, path):
    public_base_url = (
        getattr(settings, "WHATSAPP_MEDIA_BASE_URL", "")
        or getattr(settings, "EMAIL_MEDIA_BASE_URL", "")
    ).rstrip("/")
    if public_base_url:
        return f"{public_base_url}{path}"
    return request.build_absolute_uri(path)


def _build_whatsapp_share_page_url(request, attendee):
    path = reverse("attendees:whatsapp_share", args=[attendee.qr_code])
    return _build_public_absolute_url(request, path)


def _build_whatsapp_file_url(request, name, qr_code):
    return _build_public_absolute_url(request, reverse(name, args=[qr_code]))


def _build_flyer_share_payload(attendee):
    flyer_field = resolve_field_file(attendee.event, "flyer", "event_flyer") or getattr(attendee.event, "flyer", None)
    if not flyer_field:
        return None
    try:
        flyer_field.open("rb")
        content = flyer_field.read()
        flyer_field.close()
        name = str(getattr(flyer_field, "name", "") or "")
        lower_name = name.lower()
        if lower_name.endswith(".webp"):
            mimetype = "image/webp"
        elif lower_name.endswith((".jpg", ".jpeg")):
            mimetype = "image/jpeg"
        elif lower_name.endswith(".png"):
            mimetype = "image/png"
        else:
            mimetype = "application/octet-stream"
        filename = name.rsplit("/", 1)[-1] if name else f"{attendee.event.slug or attendee.event.pk}-flyer.webp"
        return {"content": content, "mimetype": mimetype, "filename": filename}
    except (FileNotFoundError, ValueError):
        return None


def _build_post_create_notice(request, branch, event):
    attendee_id = request.GET.get("created_attendee")
    email_status = request.GET.get("email_status")
    if not attendee_id or email_status != "failed":
        return None

    attendee = _attendee_queryset(branch, event).filter(pk=attendee_id).first()
    if not attendee:
        return None

    title = "Correo no enviado"
    body = (
        f"No se pudo enviar el correo a {attendee.email or 'este asistente'}. "
        "Puedes compartir la informacion por WhatsApp."
    )

    return {
        "title": title,
        "body": body,
        "attendee_name": attendee.name,
        "phone": _format_whatsapp_phone_display(attendee.phone),
        "whatsapp_url": _build_whatsapp_url(request, attendee),
        "whatsapp_share_text": build_event_share_text(
            attendee.event,
            attendee,
            qr_url=_build_whatsapp_file_url(request, "attendees:whatsapp_qr_file", attendee.qr_code),
        ),
        "whatsapp_public_qr_url": _build_whatsapp_file_url(request, "attendees:whatsapp_qr_file", attendee.qr_code),
        "whatsapp_card_url": _build_whatsapp_file_url(request, "attendees:whatsapp_card", attendee.qr_code),
        "whatsapp_qr_file_url": _build_whatsapp_file_url(request, "attendees:whatsapp_qr_file", attendee.qr_code),
        "whatsapp_flyer_file_url": _build_whatsapp_file_url(request, "attendees:whatsapp_flyer_file", attendee.qr_code),
    }


def _category_summary(branch, event):
    categories = Category.objects.filter(branch=branch, is_active=True).annotate(
        total=Count("attendees", filter=Q(attendees__branch=branch, attendees__event=event)),
        ingresados=Count(
            "attendees",
            filter=Q(attendees__branch=branch, attendees__event=event, attendees__has_checked_in=True),
        ),
        subtotal=Sum("attendees__paid_amount", filter=Q(attendees__branch=branch, attendees__event=event)),
        balance=Sum("attendees__included_balance", filter=Q(attendees__branch=branch, attendees__event=event)),
    )
    summary = []
    for category in categories:
        category.pendientes = (category.total or 0) - (category.ingresados or 0)
        category.subtotal = category.subtotal or Decimal("0")
        category.balance = category.balance or 0
        category.progress = int(((category.ingresados or 0) / category.total) * 100) if category.total else 0
        summary.append(category)
    return summary


def _list_context(request, branch, event):
    attendees = _attendee_queryset(branch, event)
    search = request.GET.get("buscar", request.GET.get("q", "")).strip()
    status = request.GET.get("estado", "").strip()

    if search:
        attendees = attendees.filter(Q(name__icontains=search) | Q(cc__icontains=search))
    if status == "ingresados":
        attendees = attendees.filter(has_checked_in=True)
    elif status == "pendientes":
        attendees = attendees.filter(has_checked_in=False)

    items_per_page = request.GET.get("items", "10")
    try:
        items_per_page = int(items_per_page)
    except (TypeError, ValueError):
        items_per_page = 10
    if items_per_page not in {10, 25, 50, 100}:
        items_per_page = 10

    paginator = Paginator(attendees, items_per_page)
    page = request.GET.get("page", 1)

    try:
        page_obj = paginator.page(page)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    for attendee in page_obj.object_list:
        attendee.whatsapp_url = _build_whatsapp_url(request, attendee)
        attendee.whatsapp_share_text = build_event_share_text(
            attendee.event,
            attendee,
            qr_url=_build_whatsapp_file_url(request, "attendees:whatsapp_qr_file", attendee.qr_code),
        )
        attendee.whatsapp_card_url = _build_whatsapp_file_url(request, "attendees:whatsapp_card", attendee.qr_code)
        attendee.whatsapp_qr_file_url = _build_whatsapp_file_url(request, "attendees:whatsapp_qr_file", attendee.qr_code)
        attendee.whatsapp_flyer_file_url = _build_whatsapp_file_url(request, "attendees:whatsapp_flyer_file", attendee.qr_code)

    return {
        "attendees": page_obj,
        "page_obj": page_obj,
        "paginator": paginator,
        "search": search,
        "status_filter": status,
        "items_per_page": items_per_page,
        "total": paginator.count,
        "showing_start": page_obj.start_index() if paginator.count else 0,
        "showing_end": page_obj.end_index() if paginator.count else 0,
    }


def _dashboard_context(
    request,
    branch,
    event,
    form=None,
    initial_tab="scanner",
    open_modal="",
    selected_category_id=None,
    category_form=None,
    editing_category=None,
    expense_form=None,
    cash_drop_form=None,
    editing_expense=None,
    editing_cash_drop=None,
):
    attendees = _attendee_queryset(branch, event)
    today = timezone.localdate()
    categories = _category_summary(branch, event)
    cash_movements = CashMovement.objects.filter(branch=branch, event=event)
    entrance_expense_movements = (
        cash_movements.filter(module=CashMovement.MODULE_ENTRANCE, movement_type=CashMovement.TYPE_EXPENSE)
        .select_related("created_by")
        .prefetch_related("payments")
    )[:10]
    entrance_cash_drop_movements = (
        cash_movements.filter(module=CashMovement.MODULE_ENTRANCE, movement_type=CashMovement.TYPE_CASH_DROP)
        .select_related("created_by")
        .prefetch_related("payments")
    )[:10]
    manual_paid = attendees.filter(origin=Attendee.ORIGIN_MANUAL).aggregate(total=Sum("paid_amount")).get("total") or Decimal("0")
    total_balance = attendees.aggregate(total=Sum("included_balance")).get("total") or 0
    total_attendees = attendees.count()
    checked_in = attendees.filter(has_checked_in=True).count()
    pending = attendees.filter(has_checked_in=False).count()
    expense_total = cash_movements.filter(movement_type=CashMovement.TYPE_EXPENSE).aggregate(total=Sum("total_amount")).get("total") or Decimal("0")
    event_day_total = cash_movements.filter(movement_type=CashMovement.TYPE_EVENT_DAY).aggregate(total=Sum("total_amount")).get("total") or Decimal("0")
    cash_drop_total = cash_movements.filter(movement_type=CashMovement.TYPE_CASH_DROP).aggregate(total=Sum("total_amount")).get("total") or Decimal("0")
    total_paid = manual_paid + event_day_total
    net_total = total_paid - expense_total

    payment_methods_raw = list(
        CashMovementPayment.objects.filter(movement__branch=branch, movement__event=event)
        .values("method")
        .annotate(total=Sum("amount"), movements=Count("id"))
        .order_by("-total")
    )
    payment_labels = dict(CashMovementPayment.METHOD_CHOICES)
    payment_total = sum((item["total"] or Decimal("0")) for item in payment_methods_raw) or Decimal("0")
    payment_methods = []
    for item in payment_methods_raw:
        method_total = item["total"] or Decimal("0")
        share = int((method_total / payment_total) * 100) if payment_total else 0
        payment_methods.append(
            {
                "key": item["method"],
                "label": payment_labels.get(item["method"], item["method"].title()),
                "total": method_total,
                "movements": item["movements"],
                "share": share,
            }
        )

    category_circles = []
    for category in categories:
        share = int(((category.total or 0) / total_attendees) * 100) if total_attendees else 0
        category_circles.append(
            {
                "name": category.name,
                "total": category.total or 0,
                "ingresados": category.ingresados or 0,
                "pendientes": category.pendientes or 0,
                "subtotal": category.subtotal or Decimal("0"),
                "progress": category.progress,
                "share": share,
            }
        )

    stats = {
        "total_asistentes": total_attendees,
        "asistentes_ingresados": checked_in,
        "pendientes": pending,
        "mis_verificaciones": attendees.filter(checked_in_by=request.user, checked_in_at__date=today).count(),
        "balance_total": total_balance,
        "total_recaudado": total_paid,
        "total_gastos": expense_total,
        "total_evento_dia": event_day_total,
        "total_vaciado": cash_drop_total,
        "total_neto": net_total,
        "progress": int((checked_in / total_attendees) * 100) if total_attendees else 0,
        "pending_progress": int((pending / total_attendees) * 100) if total_attendees else 0,
    }
    price_map = {str(category.pk): str(category.price or 0) for category in categories}

    context = {
        "branch": branch,
        "event": event,
        "form": form or _build_attendee_form(branch, event, selected_category_id=selected_category_id),
        "category_form": category_form or BranchCategoryForm(branch=branch, instance=editing_category),
        "editing_category": editing_category,
        "branch_categories": branch.categories.order_by("name"),
        "event_day_form": EventDayEntryForm(branch=branch),
        "expense_form": expense_form or _build_expense_form(editing_expense),
        "cash_drop_form": cash_drop_form or _build_cash_drop_form(editing_cash_drop),
        "categorias": categories,
        "categoria_precios_json": json.dumps(price_map),
        "stats": stats,
        "category_circles": category_circles,
        "payment_methods": payment_methods,
        "payment_total": payment_total,
        "total_recaudado": total_paid,
        "total_balance": total_balance,
        "registros": attendees.count(),
        "initial_tab": initial_tab,
        "open_modal": open_modal,
        "editing_expense": editing_expense,
        "editing_cash_drop": editing_cash_drop,
        "expense_movements": entrance_expense_movements,
        "cash_drop_movements": entrance_cash_drop_movements,
        "post_create_notice": _build_post_create_notice(request, branch, event),
    }
    context.update(_list_context(request, branch, event))
    return context


@login_required
def attendee_list(request):
    branch, event = _get_branch_and_event(request)
    if not branch or not event:
        return redirect("shared_ui:dashboard")
    if not _ensure_attendee_access(request, branch, event):
        return redirect("shared_ui:dashboard")

    requested_tab = request.GET.get("tab", "scanner")
    initial_tab = _sanitize_attendees_content_tab(requested_tab)
    open_modal = _sanitize_attendees_modal_tab(request.GET.get("modal"))
    if not open_modal and requested_tab in ATTENDEES_MODAL_TABS:
        open_modal = requested_tab
    if open_modal == "categorias" and not user_can_manage_categories(request.user, branch, event):
        messages.error(request, "Solo los administradores pueden gestionar categorias.")
        open_modal = ""
    selected_category_id = request.GET.get("selected_category")
    editing_category = None
    editing_expense = None
    editing_cash_drop = None
    if user_can_manage_categories(request.user, branch, event):
        editing_category = _get_editing_category(branch, request.GET.get("edit_category"))
    if user_can_manage_events(request.user, branch, event):
        editing_expense = _get_editing_cash_movement(
            branch,
            event,
            CashMovement.MODULE_ENTRANCE,
            CashMovement.TYPE_EXPENSE,
            request.GET.get("edit_expense"),
        )
        editing_cash_drop = _get_editing_cash_movement(
            branch,
            event,
            CashMovement.MODULE_ENTRANCE,
            CashMovement.TYPE_CASH_DROP,
            request.GET.get("edit_cash_drop"),
        )
    context = _dashboard_context(
        request,
        branch,
        event,
        initial_tab=initial_tab,
        open_modal=open_modal,
        selected_category_id=selected_category_id,
        editing_category=editing_category,
        editing_expense=editing_expense,
        editing_cash_drop=editing_cash_drop,
    )
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return render(request, "attendees/_table.html", context)
    return render(request, "attendees/list.html", context)


@login_required
def attendee_create(request):
    branch, event = _get_branch_and_event(request)
    if not branch or not event:
        return redirect("shared_ui:dashboard")
    if not _ensure_attendee_access(request, branch, event):
        return redirect("shared_ui:dashboard")

    form = AttendeeForm(request.POST or None, branch=branch, event=event)
    if request.method == "POST":
        if form.is_valid():
            attendee = form.save(commit=False)
            attendee.branch = branch
            attendee.event = event
            attendee.created_by = request.user
            attendee.included_balance = attendee.category.included_consumptions
            if not attendee.paid_amount:
                attendee.paid_amount = attendee.category.price
            attendee.save()
            email_sent = False
            email_error = None
            try:
                email_sent, email_error = send_attendee_ticket_email(attendee)
            except Exception as exc:
                email_error = str(exc)

            if email_sent:
                messages.success(
                    request,
                    f"Asistente {attendee.name} registrado correctamente.",
                )
                email_status = "accepted"
            else:
                messages.warning(
                    request,
                    f"Asistente {attendee.name} registrado, pero no se pudo enviar el correo. {email_error or ''}".strip(),
                )
                email_status = "failed"
            return redirect(
                f"{reverse('attendees:list')}?tab=crear&created_attendee={attendee.pk}&email_status={email_status}"
            )

        messages.error(request, "Corrige los datos del asistente.")
        context = _dashboard_context(request, branch, event, form=form, initial_tab="crear")
        return render(request, "attendees/list.html", context, status=400)

    return redirect(f"{reverse('attendees:list')}?tab=crear")


@login_required
def attendee_category_create(request):
    branch, event = _get_branch_and_event(request)
    if not branch or not event:
        return redirect("shared_ui:dashboard")
    if not _ensure_category_management_access(request, branch, event):
        return redirect("shared_ui:dashboard")

    form = BranchCategoryForm(request.POST or None, branch=branch)
    return_tab = _sanitize_attendees_return_tab(request.POST.get("return_tab"), default="crear")
    attendee_form = _build_attendee_form(branch, event)
    if request.method == "POST":
        if form.is_valid():
            category = form.save()
            messages.success(request, f"Categoria {category.name} creada.")
            return redirect(
                f"{reverse('attendees:list')}?tab={return_tab}&selected_category={category.pk}"
            )

        messages.error(request, "Corrige los datos de la categoria.")
        context = _dashboard_context(
            request,
            branch,
            event,
            form=attendee_form,
            initial_tab=return_tab,
            open_modal="categorias",
            category_form=form,
        )
        return render(request, "attendees/list.html", context, status=400)

    return redirect("attendees:list")


@login_required
def attendee_category_update(request, category_id):
    branch, event = _get_branch_and_event(request)
    if not branch or not event:
        return redirect("shared_ui:dashboard")
    if not _ensure_category_management_access(request, branch, event):
        return redirect("shared_ui:dashboard")

    category = get_object_or_404(Category, pk=category_id, branch=branch)
    form = BranchCategoryForm(request.POST or None, branch=branch, instance=category)
    return_tab = _sanitize_attendees_return_tab(request.POST.get("return_tab"), default="crear")
    attendee_form = _build_attendee_form(branch, event)
    if request.method == "POST":
        if form.is_valid():
            category = form.save()
            messages.success(request, f"Categoria {category.name} actualizada.")
            return redirect(
                f"{reverse('attendees:list')}?tab={return_tab}&selected_category={category.pk}"
            )

        messages.error(request, "Corrige los datos de la categoria.")
        context = _dashboard_context(
            request,
            branch,
            event,
            form=attendee_form,
            initial_tab=return_tab,
            open_modal="categorias",
            category_form=form,
            editing_category=category,
        )
        return render(request, "attendees/list.html", context, status=400)

    return redirect(f"{reverse('attendees:list')}?tab={return_tab}&modal=categorias&edit_category={category.pk}")


@require_POST
@login_required
def attendee_category_delete(request, category_id):
    branch, event = _get_branch_and_event(request)
    if not branch or not event:
        return redirect("shared_ui:dashboard")
    if not _ensure_category_management_access(request, branch, event):
        return redirect("shared_ui:dashboard")

    category = get_object_or_404(Category, pk=category_id, branch=branch)
    return_tab = _sanitize_attendees_return_tab(request.POST.get("return_tab"), default="crear")
    category_name = category.name
    result = delete_branch_category(category)
    if result == "deleted":
        messages.success(request, f"Categoria {category_name} eliminada.")
    elif result == "deactivated":
        messages.warning(
            request,
            f"Categoria {category_name} inactivada porque ya tiene asistentes asociados.",
        )
    else:
        messages.info(request, f"La categoria {category_name} ya estaba inactiva.")

    return redirect(f"{reverse('attendees:list')}?tab={return_tab}&modal=categorias")


@login_required
def attendee_event_day_create(request):
    branch, event = _get_branch_and_event(request)
    if not branch or not event:
        return redirect("shared_ui:dashboard")
    if not _ensure_attendee_access(request, branch, event):
        return redirect("shared_ui:dashboard")

    attendee_form = AttendeeForm(branch=branch, event=event)
    form = EventDayEntryForm(request.POST or None, branch=branch)
    if request.method == "POST":
        try:
            payments = extract_split_payments(request.POST, request.FILES, prefix="event_day")
        except ValueError as exc:
            payments = None
            messages.error(request, str(exc))
        if form.is_valid() and payments is not None:
            try:
                movement = register_event_day_entry(
                    branch=branch,
                    event=event,
                    category=form.cleaned_data["category"],
                    attendee_quantity=form.cleaned_data["attendee_quantity"],
                    unit_amount=form.cleaned_data["unit_amount"],
                    user=request.user,
                    description=form.cleaned_data["description"],
                    payments=payments,
                )
            except ValueError as exc:
                messages.error(request, str(exc))
            else:
                messages.success(
                    request,
                    f"Dia de evento registrado: {movement.attendee_quantity} asistentes por $ {movement.unit_amount}."
                )
                return redirect(f"{reverse('attendees:list')}?tab=lista")

        context = _dashboard_context(request, branch, event, form=attendee_form, initial_tab="evento-dia")
        context["event_day_form"] = form
        return render(request, "attendees/list.html", context, status=400)

    return redirect("attendees:list")


@login_required
def attendee_expense_create(request):
    branch, event = _get_branch_and_event(request)
    if not branch or not event:
        return redirect("shared_ui:dashboard")
    if not _ensure_attendee_access(request, branch, event):
        return redirect("shared_ui:dashboard")

    attendee_form = AttendeeForm(branch=branch, event=event)
    form = ExpenseForm(request.POST or None, request.FILES or None)
    if request.method == "POST":
        if form.is_valid():
            try:
                payments = resolve_expense_payments(request.POST, request.FILES, form, prefix="expense")
                create_cash_movement(
                    branch=branch,
                    event=event,
                    user=request.user,
                    module=CashMovement.MODULE_ENTRANCE,
                    movement_type=CashMovement.TYPE_EXPENSE,
                    total_amount=form.cleaned_data["amount"],
                    description=form.cleaned_data["description"],
                    payments=payments,
                )
            except ValueError as exc:
                messages.error(request, str(exc))
            else:
                messages.success(request, "Gasto registrado en entrada.")
                return redirect(f"{reverse('attendees:list')}?tab=lista")
        else:
            messages.error(request, "Corrige los datos del gasto.")

        context = _dashboard_context(request, branch, event, form=attendee_form, initial_tab="gastos")
        context["expense_form"] = form
        return render(request, "attendees/list.html", context, status=400)

    return redirect("attendees:list")


@require_POST
@login_required
def attendee_expense_update(request, movement_id):
    branch, event = _get_branch_and_event(request)
    if not branch or not event:
        return redirect("shared_ui:dashboard")
    if not _ensure_cash_management_access(request, branch, event):
        return redirect("shared_ui:dashboard")

    movement = get_object_or_404(
        CashMovement,
        pk=movement_id,
        branch=branch,
        event=event,
        module=CashMovement.MODULE_ENTRANCE,
        movement_type=CashMovement.TYPE_EXPENSE,
    )
    attendee_form = AttendeeForm(branch=branch, event=event)
    form = ExpenseForm(request.POST or None, request.FILES or None)
    if request.method == "POST":
        if form.is_valid():
            payments = None
            if _expense_payment_inputs_present(request, prefix="expense"):
                try:
                    payments = resolve_expense_payments(request.POST, request.FILES, form, prefix="expense")
                except ValueError as exc:
                    messages.error(request, str(exc))
                    context = _dashboard_context(
                        request,
                        branch,
                        event,
                        form=attendee_form,
                        initial_tab="gastos",
                        expense_form=form,
                        editing_expense=movement,
                    )
                    return render(request, "attendees/list.html", context, status=400)
            try:
                update_cash_movement(
                    movement=movement,
                    total_amount=form.cleaned_data["amount"],
                    description=form.cleaned_data["description"],
                    payments=payments,
                )
            except ValueError as exc:
                messages.error(request, str(exc))
                return redirect(f"{reverse('attendees:list')}?tab=gastos&edit_expense={movement.id}")
            messages.success(request, "Gasto actualizado en entrada.")
            return redirect(f"{reverse('attendees:list')}?tab=gastos")
        messages.error(request, "Corrige los datos del gasto.")

        context = _dashboard_context(
            request,
            branch,
            event,
            form=attendee_form,
            initial_tab="gastos",
            expense_form=form,
            editing_expense=movement,
        )
        return render(request, "attendees/list.html", context, status=400)

    return redirect("attendees:list")


@require_POST
@login_required
def attendee_expense_delete(request, movement_id):
    branch, event = _get_branch_and_event(request)
    if not branch or not event:
        return redirect("shared_ui:dashboard")
    if not _ensure_cash_management_access(request, branch, event):
        return redirect("shared_ui:dashboard")

    movement = get_object_or_404(
        CashMovement,
        pk=movement_id,
        branch=branch,
        event=event,
        module=CashMovement.MODULE_ENTRANCE,
        movement_type=CashMovement.TYPE_EXPENSE,
    )
    delete_cash_movement(movement=movement)
    messages.success(request, "Gasto eliminado de entrada.")
    return redirect(f"{reverse('attendees:list')}?tab=gastos")


@login_required
def attendee_cash_drop_create(request):
    branch, event = _get_branch_and_event(request)
    if not branch or not event:
        return redirect("shared_ui:dashboard")
    if not _ensure_attendee_access(request, branch, event):
        return redirect("shared_ui:dashboard")

    attendee_form = AttendeeForm(branch=branch, event=event)
    form = CashDropForm(request.POST or None)
    if request.method == "POST":
        if form.is_valid():
            try:
                create_cash_movement(
                    branch=branch,
                    event=event,
                    user=request.user,
                    module=CashMovement.MODULE_ENTRANCE,
                    movement_type=CashMovement.TYPE_CASH_DROP,
                    total_amount=form.cleaned_data["amount"],
                    description=form.cleaned_data["description"],
                )
            except ValueError as exc:
                messages.error(request, str(exc))
            else:
                messages.success(request, "Vaciado de caja registrado.")
                return redirect(f"{reverse('attendees:list')}?tab=lista")
        else:
            messages.error(request, "Corrige los datos del vaciado de caja.")

        context = _dashboard_context(request, branch, event, form=attendee_form, initial_tab="vaciar-caja")
        context["cash_drop_form"] = form
        return render(request, "attendees/list.html", context, status=400)

    return redirect("attendees:list")


@require_POST
@login_required
def attendee_cash_drop_update(request, movement_id):
    branch, event = _get_branch_and_event(request)
    if not branch or not event:
        return redirect("shared_ui:dashboard")
    if not _ensure_cash_management_access(request, branch, event):
        return redirect("shared_ui:dashboard")

    movement = get_object_or_404(
        CashMovement,
        pk=movement_id,
        branch=branch,
        event=event,
        module=CashMovement.MODULE_ENTRANCE,
        movement_type=CashMovement.TYPE_CASH_DROP,
    )
    attendee_form = AttendeeForm(branch=branch, event=event)
    form = CashDropForm(request.POST or None)
    if request.method == "POST":
        if form.is_valid():
            try:
                update_cash_movement(
                    movement=movement,
                    total_amount=form.cleaned_data["amount"],
                    description=form.cleaned_data["description"],
                )
            except ValueError as exc:
                messages.error(request, str(exc))
                return redirect(f"{reverse('attendees:list')}?tab=vaciar-caja&edit_cash_drop={movement.id}")
            messages.success(request, "Vaciado de caja actualizado.")
            return redirect(f"{reverse('attendees:list')}?tab=vaciar-caja")
        messages.error(request, "Corrige los datos del vaciado de caja.")

        context = _dashboard_context(
            request,
            branch,
            event,
            form=attendee_form,
            initial_tab="vaciar-caja",
            cash_drop_form=form,
            editing_cash_drop=movement,
        )
        return render(request, "attendees/list.html", context, status=400)

    return redirect("attendees:list")


@require_POST
@login_required
def attendee_cash_drop_delete(request, movement_id):
    branch, event = _get_branch_and_event(request)
    if not branch or not event:
        return redirect("shared_ui:dashboard")
    if not _ensure_cash_management_access(request, branch, event):
        return redirect("shared_ui:dashboard")

    movement = get_object_or_404(
        CashMovement,
        pk=movement_id,
        branch=branch,
        event=event,
        module=CashMovement.MODULE_ENTRANCE,
        movement_type=CashMovement.TYPE_CASH_DROP,
    )
    delete_cash_movement(movement=movement)
    messages.success(request, "Vaciado de caja eliminado.")
    return redirect(f"{reverse('attendees:list')}?tab=vaciar-caja")


@require_POST
@login_required
def attendee_check_in(request):
    branch, event = _get_branch_and_event(request)
    if not branch or not event:
        return JsonResponse({"success": False, "message": "Selecciona sucursal y evento."}, status=400)
    if not user_can_access_attendees(request.user, branch, event):
        return JsonResponse({"success": False, "message": "Sin permisos para entrada."}, status=403)

    code = request.POST.get("code", "").strip()
    attendee = get_attendee_for_branch(branch, event, code)

    if not attendee:
        return JsonResponse({"success": False, "message": "No se encontro el asistente."}, status=404)

    attendee, created = check_in_attendee(attendee, request.user)
    if not created:
        checked_in_at = timezone.localtime(attendee.checked_in_at).strftime("%d/%m/%Y %H:%M") if attendee.checked_in_at else ""
        return JsonResponse(
            {
                "success": False,
                "message": "El asistente ya ingreso.",
                "attendee": {"name": attendee.name, "cc": attendee.cc, "checked_in_at": checked_in_at},
            }
        )

    return JsonResponse(
        {
            "success": True,
            "message": "Ingreso registrado.",
            "attendee": {
                "name": attendee.name,
                "cc": attendee.cc,
                "category": attendee.category.name,
                "balance": attendee.included_balance,
            },
        }
    )


def _parse_json_request(request):
    try:
        return json.loads(request.body.decode("utf-8")) if request.body else {}
    except json.JSONDecodeError:
        return None


@require_POST
@login_required
def attendee_check_in_preview(request):
    branch, event = _get_branch_and_event(request)
    if not branch or not event:
        return JsonResponse({"success": False, "message": "Selecciona sucursal y evento."}, status=400)
    if not user_can_access_attendees(request.user, branch, event):
        return JsonResponse({"success": False, "message": "Sin permisos para entrada."}, status=403)

    payload = _parse_json_request(request)
    if payload is None:
        return JsonResponse({"success": False, "message": "Datos invalidos."}, status=400)

    code = str(payload.get("codigo") or payload.get("code") or "").strip()
    if not code:
        return JsonResponse({"success": False, "message": "Codigo QR vacio."}, status=400)

    attendee = get_attendee_for_branch(branch, event, code)
    if not attendee:
        return JsonResponse({"success": False, "message": "Codigo QR invalido o asistente no encontrado."}, status=404)

    if attendee.has_checked_in:
        checked_in_at = timezone.localtime(attendee.checked_in_at).strftime("%d/%m/%Y %H:%M") if attendee.checked_in_at else ""
        verified_by = attendee.checked_in_by.username if attendee.checked_in_by else "N/A"
        return JsonResponse(
            {
                "success": False,
                "message": f"Este asistente ya ingreso el {checked_in_at} verificado por {verified_by}.",
                "attendee": {"name": attendee.name},
            }
        )

    return JsonResponse(
        {
            "success": True,
            "attendee": {
                "name": attendee.name,
                "cc": attendee.cc,
                "phone": attendee.phone,
                "email": attendee.email,
                "category": attendee.category.name,
                "balance": attendee.included_balance,
                "paid_amount": str(attendee.paid_amount),
                "created_at": timezone.localtime(attendee.created_at).strftime("%d/%m/%Y"),
            },
        }
    )


@require_POST
@login_required
def attendee_confirm_check_in(request):
    branch, event = _get_branch_and_event(request)
    if not branch or not event:
        return JsonResponse({"success": False, "message": "Selecciona sucursal y evento."}, status=400)
    if not user_can_access_attendees(request.user, branch, event):
        return JsonResponse({"success": False, "message": "Sin permisos para entrada."}, status=403)

    payload = _parse_json_request(request)
    if payload is None:
        return JsonResponse({"success": False, "message": "Datos invalidos."}, status=400)

    code = str(payload.get("codigo") or payload.get("code") or "").strip()
    if not code:
        return JsonResponse({"success": False, "message": "Codigo QR vacio."}, status=400)

    attendee = get_attendee_for_branch(branch, event, code)
    if not attendee:
        return JsonResponse({"success": False, "message": "Asistente no encontrado."}, status=404)

    attendee, created = check_in_attendee(attendee, request.user)
    if not created:
        return JsonResponse({"success": False, "message": "El asistente ya habia ingresado."}, status=400)

    checked_in_at = timezone.localtime(attendee.checked_in_at)
    return JsonResponse(
        {
            "success": True,
            "attendee": {
                "name": attendee.name,
                "cc": attendee.cc,
                "category": attendee.category.name,
                "balance": attendee.included_balance,
                "time": checked_in_at.strftime("%H:%M:%S"),
                "date": checked_in_at.strftime("%d/%m/%Y"),
            },
        }
    )


@require_POST
@login_required
def attendee_mark_checked_in(request):
    branch, event = _get_branch_and_event(request)
    if not branch or not event:
        return JsonResponse({"success": False, "message": "Selecciona sucursal y evento."}, status=400)
    if not user_can_access_attendees(request.user, branch, event):
        return JsonResponse({"success": False, "message": "Sin permisos para entrada."}, status=403)

    payload = _parse_json_request(request)
    if payload is None:
        return JsonResponse({"success": False, "message": "Datos invalidos."}, status=400)

    cc = str(payload.get("cc") or "").strip()
    attendee = _attendee_queryset(branch, event).filter(cc=cc).first()
    if not attendee:
        return JsonResponse({"success": False, "message": "Asistente no encontrado."}, status=404)

    attendee, created = check_in_attendee(attendee, request.user)
    if not created:
        return JsonResponse({"success": False, "message": "El asistente ya habia ingresado."}, status=400)
    return JsonResponse({"success": True, "message": f"{attendee.name} marcado como ingresado."})


@require_POST
@login_required
def attendee_delete(request):
    branch, event = _get_branch_and_event(request)
    if not branch or not event:
        return JsonResponse({"success": False, "message": "Selecciona sucursal y evento."}, status=400)
    if not user_can_access_attendees(request.user, branch, event):
        return JsonResponse({"success": False, "message": "Sin permisos para entrada."}, status=403)

    payload = _parse_json_request(request)
    if payload is None:
        return JsonResponse({"success": False, "message": "Datos invalidos."}, status=400)

    cc = str(payload.get("cc") or "").strip()
    attendee = _attendee_queryset(branch, event).filter(cc=cc).first()
    if not attendee:
        return JsonResponse({"success": False, "message": "Asistente no encontrado."}, status=404)
    can_manage_checked_in_attendees = user_can_manage_events(request.user, branch, event)
    if attendee.has_checked_in and not can_manage_checked_in_attendees:
        return JsonResponse({"success": False, "message": "No se puede eliminar un asistente que ya ingreso."}, status=400)

    name = attendee.name
    attendee.delete()
    return JsonResponse({"success": True, "message": f"{name} eliminado exitosamente."})


@require_GET
@login_required
def attendee_qr_detail(request, cc):
    branch, event = _get_branch_and_event(request)
    if not branch or not event:
        return JsonResponse({"success": False, "message": "Selecciona sucursal y evento."}, status=400)
    if not user_can_access_attendees(request.user, branch, event):
        return JsonResponse({"success": False, "message": "Sin permisos para entrada."}, status=403)

    attendee = get_object_or_404(_attendee_queryset(branch, event), cc=cc)
    return JsonResponse(
        {
            "success": True,
            "qr_url": attendee.qr_image.url if attendee.qr_image else "",
            "attendee": {
                "name": attendee.name,
                "cc": attendee.cc,
                "category": attendee.category.name,
                "balance": attendee.included_balance,
                "paid_amount": str(attendee.paid_amount),
            },
        }
    )


@require_GET
def attendee_whatsapp_share(request, qr_code):
    attendee = get_object_or_404(
        Attendee.objects.select_related("branch", "event", "category"),
        qr_code=qr_code,
    )
    card_url = _build_public_absolute_url(request, reverse("attendees:whatsapp_card", args=[attendee.qr_code]))
    context = {
        "attendee": attendee,
        "share_title": f"{attendee.event.name} - {attendee.name}",
        "share_description": "Acceso del evento con QR y datos del asistente.",
        "card_url": card_url,
    }
    return render(request, "attendees/whatsapp_share.html", context)


@require_GET
def attendee_whatsapp_card(request, qr_code):
    attendee = get_object_or_404(
        Attendee.objects.select_related("branch", "event", "category"),
        qr_code=qr_code,
    )
    image_bytes = build_whatsapp_share_card_png(attendee)
    return HttpResponse(image_bytes, content_type="image/png")


@require_GET
@login_required
def attendee_whatsapp_qr_file(request, qr_code):
    branch, event = _get_branch_and_event(request)
    if not branch or not event:
        return redirect("shared_ui:dashboard")
    if not _ensure_attendee_access(request, branch, event):
        return redirect("shared_ui:dashboard")

    attendee = get_object_or_404(_attendee_queryset(branch, event), qr_code=qr_code)
    image_bytes = build_qr_png_bytes(attendee.qr_code, attendee.event, attendee.branch)
    response = HttpResponse(image_bytes, content_type="image/png")
    response["Content-Disposition"] = f'inline; filename="{attendee.qr_code}.png"'
    return response


@require_GET
@login_required
def attendee_whatsapp_flyer_file(request, qr_code):
    branch, event = _get_branch_and_event(request)
    if not branch or not event:
        return redirect("shared_ui:dashboard")
    if not _ensure_attendee_access(request, branch, event):
        return redirect("shared_ui:dashboard")

    attendee = get_object_or_404(_attendee_queryset(branch, event), qr_code=qr_code)
    payload = _build_flyer_share_payload(attendee)
    if not payload:
        return JsonResponse({"success": False, "message": "El evento no tiene flyer configurado."}, status=404)

    response = HttpResponse(payload["content"], content_type=payload["mimetype"])
    response["Content-Disposition"] = f'inline; filename="{payload["filename"]}"'
    return response


@require_GET
@login_required
def attendee_export_excel(request):
    branch, event = _get_branch_and_event(request)
    if not branch or not event:
        return redirect("shared_ui:dashboard")
    if not _ensure_attendee_access(request, branch, event):
        return redirect("shared_ui:dashboard")

    attendees = _attendee_queryset(branch, event)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = f"Entrada {event.name[:18]}"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0F5132", end_color="0F5132", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")

    headers = [
        "NOMBRE",
        "CEDULA",
        "TELEFONO",
        "CORREO",
        "CATEGORIA",
        "PRECIO PAGADO",
        "BALANCE",
        "ESTADO",
        "FECHA REGISTRO",
        "FECHA INGRESO",
        "VERIFICADO POR",
    ]

    for index, value in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=index, value=value)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for row_index, attendee in enumerate(attendees, start=2):
        checked_in_at = timezone.localtime(attendee.checked_in_at) if attendee.checked_in_at else None
        sheet.cell(row=row_index, column=1, value=attendee.name)
        sheet.cell(row=row_index, column=2, value=attendee.cc)
        sheet.cell(row=row_index, column=3, value=attendee.phone or "")
        sheet.cell(row=row_index, column=4, value=attendee.email or "")
        sheet.cell(row=row_index, column=5, value=attendee.category.name)
        sheet.cell(row=row_index, column=6, value=float(attendee.paid_amount or 0))
        sheet.cell(row=row_index, column=7, value=attendee.included_balance)
        sheet.cell(row=row_index, column=8, value="INGRESO" if attendee.has_checked_in else "PENDIENTE")
        sheet.cell(row=row_index, column=9, value=timezone.localtime(attendee.created_at).strftime("%d/%m/%Y %H:%M"))
        sheet.cell(row=row_index, column=10, value=checked_in_at.strftime("%d/%m/%Y %H:%M") if checked_in_at else "")
        sheet.cell(row=row_index, column=11, value=attendee.checked_in_by.username if attendee.checked_in_by else "")

    widths = [28, 18, 16, 28, 18, 14, 12, 12, 20, 20, 18]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(index)].width = width

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"entrada_{branch.slug}_{event.slug}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response
